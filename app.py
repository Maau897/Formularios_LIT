from __future__ import annotations

from calendar import monthrange
from dataclasses import asdict, dataclass
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path
import json
import os
import re
import shutil
import time
from typing import Any
import unicodedata
from uuid import uuid4
from zoneinfo import ZoneInfo

import pandas as pd
import streamlit as st
try:
    from dotenv import load_dotenv
except ModuleNotFoundError:
    def load_dotenv() -> bool:
        return False
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU
from PIL import Image as PILImage
from microsoft_graph import (
    MicrosoftGraphError,
    add_master_table_row,
    configure_microsoft_graph,
    get_master_table,
    microsoft_graph_enabled,
    update_master_table_row,
)
from supabase_users import (
    aprobar_usuario,
    actualizar_nombre_usuario,
    actualizar_rol_usuario,
    autenticar_usuario,
    configure_supabase_users,
    crear_admin_inicial,
    eliminar_usuario,
    listar_eventos_auditoria,
    listar_usuarios,
    registrar_evento_auditoria,
    obtener_usuarios_pendientes,
    registrar_usuario,
    sincronizar_usuarios_conocidos,
    supabase_users_enabled,
)
from supabase_storage import (
    configure_supabase_storage,
    download_template_bytes,
    download_signature_bytes,
    get_signatures_storage_cache_key,
    get_templates_storage_cache_key,
    load_equipment_config_payload as load_remote_equipment_config_payload,
    list_signature_assets as list_remote_signature_assets,
    list_periods as list_remote_periods,
    list_traceability_entries as list_remote_traceability_entries,
    load_period_payload as load_remote_period_payload,
    save_equipment_config_payload as save_remote_equipment_config_payload,
    save_period_payload as save_remote_period_payload,
    save_traceability_entry as save_remote_traceability_entry,
    delete_traceability_entry as delete_remote_traceability_entry,
    signatures_storage_enabled,
    supabase_storage_enabled,
    templates_storage_enabled,
)


load_dotenv()

BASE_DIR = Path(__file__).resolve().parent
TEMPLATE_PATH = BASE_DIR / "F-LIT-21-03.xlsx"
WORKING_TEMPLATE_PATH = BASE_DIR / "template_cong1.xlsx"
DATA_DIR = BASE_DIR / "data"
TRACEABILITY_DIR = DATA_DIR / "trazabilidad"
EQUIPMENT_CONFIG_DIR = DATA_DIR / "configuracion_equipos"
SIGNATURES_DIR = BASE_DIR / "firmas digitales"

MONTHS = {
    1: "ENERO",
    2: "FEBRERO",
    3: "MARZO",
    4: "ABRIL",
    5: "MAYO",
    6: "JUNIO",
    7: "JULIO",
    8: "AGOSTO",
    9: "SEPTIEMBRE",
    10: "OCTUBRE",
    11: "NOVIEMBRE",
    12: "DICIEMBRE",
}

TIME_SLOTS = [
    "7:00 - 10:00",
    "11:00 - 14:00",
    "15:00 - 18:00",
]

TRACEABILITY_TYPES = {
    "mantenimiento": "Mantenimiento",
    "calibracion": "Calibracion",
    "calificacion": "Calificacion",
    "validacion_aplicacion": "Validacion de la aplicacion",
}

TRACEABILITY_STATUSES = {
    "programado": "Programado",
    "realizado": "Realizado",
    "no_aplica": "No aplica",
}

DAY_BLOCK_START_COLUMNS = {
    day: 5 + ((day - 1) * 3)
    for day in range(1, 17)
}
DAY_BLOCK_START_COLUMNS.update(
    {day: 5 + ((day - 17) * 3) for day in range(17, 32)}
)

ROW_GROUPS = {
    "top": {"temperature": 16, "performed_by": 19, "verified_by": 20, "date": 21},
    "bottom": {"temperature": 28, "performed_by": 31, "verified_by": 32, "date": 33},
}

CORRECTION_CELL_MAP = {
    "range_1": "AU7",
    "range_2": "AW7",
    "range_3": "AY7",
}

HEADER_CELL_MAP = {
    "month": "P6",
    "year": "X6",
}

FOOTER_CELL_MAP = {
    "observations": "H34",
    "reviewed_by": "H35",
    "reviewed_on": "AI35",
}

FORM_DEFINITIONS: dict[str, dict[str, Any]] = {
    "congeladores": {
        "label": "F-LIT-21-03 Congeladores",
        "source_file": "F-LIT-21-03.xlsx",
        "working_file": "template_cong1.xlsx",
        "sheet_exclusions": {"CONG"},
        "default_equipment": "CONG-1",
        "supports_corrections": True,
        "metrics": [
            {"key": "measured_temperatures", "label": "Temperatura medida", "unit": "°C", "corrected": True},
        ],
        "layout": {
            "top": {"metric_1": 16, "performed_by": 19, "verified_by": 20, "date": 21},
            "bottom": {"metric_1": 28, "performed_by": 31, "verified_by": 32, "date": 33},
            "footer": {"observations": "H34", "reviewed_by": "H35", "reviewed_on": "AI35"},
            "status_rows_to_merge": [20, 21, 32, 33],
        },
        "extractor": "cold_equipment",
    },
    "ultracongeladores": {
        "label": "F-LIT-20-03 Ultracongeladores",
        "source_file": "F-LIT-20-03 ulcos.xlsx",
        "working_file": "template_ulcos.xlsx",
        "sheet_exclusions": {"ULCO"},
        "default_equipment": "ULCO-1",
        "supports_corrections": True,
        "metrics": [
            {"key": "measured_temperatures", "label": "Temperatura medida", "unit": "°C", "corrected": True},
        ],
        "layout": {
            "top": {"metric_1": 16, "performed_by": 19, "verified_by": 20, "date": 21},
            "bottom": {"metric_1": 28, "performed_by": 31, "verified_by": 32, "date": 33},
            "footer": {"observations": "H34", "reviewed_by": "H35", "reviewed_on": "AI35"},
            "status_rows_to_merge": [20, 21, 32, 33],
        },
        "extractor": "cold_equipment",
    },
    "refrigeradores": {
        "label": "F-LIT-22-03 Refrigeradores",
        "source_file": "F-LIT-22-03 regrigeradores.xlsx",
        "working_file": "template_refrigeradores.xlsx",
        "sheet_exclusions": {"REFR"},
        "default_equipment": "REFR-1",
        "supports_corrections": True,
        "metrics": [
            {"key": "measured_temperatures", "label": "Temperatura medida", "unit": "°C", "corrected": True},
        ],
        "layout": {
            "top": {"metric_1": 16, "performed_by": 19, "verified_by": 20, "date": 21},
            "bottom": {"metric_1": 28, "performed_by": 31, "verified_by": 32, "date": 33},
            "footer": {"observations": "H34", "reviewed_by": "H35", "reviewed_on": "AI35"},
            "status_rows_to_merge": [20, 21, 32, 33],
        },
        "extractor": "cold_equipment",
    },
    "incubadoras": {
        "label": "F-LIT-23-03 Incubadoras",
        "source_file": "F-LIT-23-03 incubadoras.xlsx",
        "working_file": "template_incubadoras.xlsx",
        "sheet_exclusions": {"ICO2"},
        "default_equipment": "ICO2-1",
        "supports_corrections": True,
        "metrics": [
            {"key": "measured_temperatures", "label": "Temperatura medida", "unit": "°C", "corrected": True},
            {"key": "secondary_measurements", "label": "%CO2", "unit": "", "corrected": False},
        ],
        "layout": {
            "top": {"metric_1": 16, "metric_2": 19, "performed_by": 22, "verified_by": 23, "date": 24},
            "bottom": {"metric_1": 31, "metric_2": 34, "performed_by": 37, "verified_by": 38, "date": 39},
            "footer": {"observations": "K40", "reviewed_by": "K41", "reviewed_on": "AL41"},
            "status_rows_to_merge": [23, 24, 38, 39],
        },
        "extractor": "incubators",
    },
    "condiciones_ambientales": {
        "label": "F-LIT-09-04 Condiciones Ambientales",
        "source_file": "F-LIT-09-04 condiciones ambientales.xlsx",
        "working_file": "template_condiciones_ambientales.xlsx",
        "sheet_exclusions": {"TEMPERATURA", "HUMEDAD"},
        "default_equipment": "TEMPERATURA 504",
        "supports_corrections": True,
        "metrics": [
            {"key": "measured_temperatures", "label": "Lectura medida", "unit": "", "corrected": True},
        ],
        "layout": {
            "top": {"metric_1": 17, "performed_by": 20, "verified_by": 21, "date": 22},
            "bottom": {"metric_1": 29, "performed_by": 32, "verified_by": 33, "date": 34},
            "footer": {"observations": "E35", "reviewed_by": "E36", "reviewed_on": "AE36"},
            "status_rows_to_merge": [21, 22, 33, 34],
        },
        "extractor": "ambient",
    },
}

DEFAULT_FORM_KEY = "congeladores"
DEFAULT_EQUIPMENT_CODE = FORM_DEFINITIONS[DEFAULT_FORM_KEY]["default_equipment"]
EQUIPMENT_CONFIG_CACHE_VERSION = "2026-06-03-admin-range-editing"
TEMPERATURE_DECIMAL_PLACES = 3
ROLES_USUARIO = ["captura", "calidad", "admin"]
LEGACY_ROLE_MAP = {
    "responsable": "admin",
    "auditor": "calidad",
}
SENSITIVE_EDITOR_ROLES = {"admin"}
KNOWN_USER_DISPLAY_NAMES = {
    "itzbloodcor@gmail.com": "Itzel",
    "miltoonnietoo.66@gmail.com": "Milton",
    "mercedesviettri@gmail.com": "Mercedes",
    "drhzamudio@gmail.com": "Horacio",
    "rodolfo_chvz@outlook.com": "Rodolfo",
    "helios.avel@gmail.com": "Angelica",
}
AUTOSAVE_DEBOUNCE_SECONDS = 3.0
LOCAL_TIMEZONE = ZoneInfo("America/Mexico_City")


@dataclass
class DailyCapture:
    active: bool
    measured_temperatures: list[str]
    corrected_temperatures: list[str]
    secondary_measurements: list[str]
    performed_by_slots: list[str]
    canceled_slots: list[int]
    verified_by: str
    recorded_on: str
    recorded_on_mode: str = "auto"
    notes: str = ""
    cancellation_note: str = ""


def default_daily_capture(day: int) -> DailyCapture:
    is_weekday = date.today().replace(day=min(day, 28)).weekday() < 5
    return DailyCapture(
        active=is_weekday,
        measured_temperatures=["", "", ""],
        corrected_temperatures=["", "", ""],
        secondary_measurements=["", "", ""],
        performed_by_slots=["", "", ""],
        canceled_slots=[],
        cancellation_note="",
        verified_by="",
        recorded_on="",
        recorded_on_mode="auto",
        notes="",
    )


def get_config_value(secret_key: str, env_key: str, default: Any = "") -> Any:
    try:
        return st.secrets.get(secret_key, os.getenv(env_key, default))
    except Exception:
        return os.getenv(env_key, default)


def as_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.strip().lower() in {"1", "true", "yes", "si", "on"}
    return bool(value)


def configure_users_backend() -> None:
    configure_supabase_users(
        url=str(get_config_value("supabase_url", "SUPABASE_URL", "")),
        key=str(get_config_value("supabase_key", "SUPABASE_KEY", "")),
        enabled=as_bool(get_config_value("use_supabase_users", "USE_SUPABASE_USERS", False)),
        table_name=str(get_config_value("supabase_users_table", "SUPABASE_USERS_TABLE", "usuarios_app")),
        audit_table_name=str(get_config_value("supabase_audit_table", "SUPABASE_AUDIT_TABLE", "formularios_auditoria")),
    )

    admin_email = str(get_config_value("admin_email", "ADMIN_EMAIL", "")).strip()
    admin_password = str(get_config_value("admin_password", "ADMIN_PASSWORD", "")).strip()
    if supabase_users_enabled() and admin_email and admin_password:
        try:
            crear_admin_inicial(admin_email, admin_password)
        except Exception:
            pass
    if supabase_users_enabled():
        try:
            sincronizar_usuarios_conocidos(KNOWN_USER_DISPLAY_NAMES, force_admin=False)
        except Exception:
            pass


def configure_storage_backend() -> None:
    configure_supabase_storage(
        url=str(get_config_value("supabase_url", "SUPABASE_URL", "")),
        key=str(get_config_value("supabase_key", "SUPABASE_KEY", "")),
        enabled=as_bool(get_config_value("use_supabase_storage", "USE_SUPABASE_STORAGE", True)),
        table_name=str(get_config_value("supabase_storage_table", "SUPABASE_STORAGE_TABLE", "formularios_periodos")),
        signatures_bucket=str(get_config_value("supabase_signatures_bucket", "SUPABASE_SIGNATURES_BUCKET", "firmas-digitales")),
        signatures_prefix=str(get_config_value("supabase_signatures_prefix", "SUPABASE_SIGNATURES_PREFIX", "")),
        templates_bucket=str(get_config_value("supabase_templates_bucket", "SUPABASE_TEMPLATES_BUCKET", "")),
        templates_prefix=str(get_config_value("supabase_templates_prefix", "SUPABASE_TEMPLATES_PREFIX", "")),
        equipment_config_table_name=str(get_config_value("supabase_equipment_config_table", "SUPABASE_EQUIPMENT_CONFIG_TABLE", "formularios_equipo_config")),
    )


def configure_microsoft_backend() -> None:
    configure_microsoft_graph(
        tenant_id=str(get_config_value("microsoft_tenant_id", "MICROSOFT_TENANT_ID", "")),
        client_id=str(get_config_value("microsoft_client_id", "MICROSOFT_CLIENT_ID", "")),
        client_secret=str(get_config_value("microsoft_client_secret", "MICROSOFT_CLIENT_SECRET", "")),
        refresh_token=str(get_config_value("microsoft_refresh_token", "MICROSOFT_REFRESH_TOKEN", "")),
        shared_url=str(get_config_value("microsoft_shared_url", "MICROSOFT_SHARED_URL", "")),
        table_name=str(get_config_value("microsoft_table_name", "MICROSOFT_TABLE_NAME", "ListaMaestra")),
    )


def normalize_user_role(rol: Any, es_admin: bool) -> str:
    if es_admin:
        return "admin"
    normalized = str(rol or "captura").strip().lower()
    normalized = LEGACY_ROLE_MAP.get(normalized, normalized)
    return normalized if normalized in ROLES_USUARIO else "captura"


def normalize_user_display_name(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or "").strip())
    text = "".join(character for character in text if not unicodedata.combining(character))
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def get_known_display_name(email: str) -> str:
    return KNOWN_USER_DISPLAY_NAMES.get(email.strip().lower(), "")


def get_current_user_display_name() -> str:
    stored_name = normalize_user_display_name(st.session_state.get("usuario_nombre", ""))
    if stored_name:
        return stored_name
    email = str(st.session_state.get("usuario_email", "")).strip().lower()
    known_name = get_known_display_name(email)
    if known_name:
        return known_name
    local_part = email.split("@", 1)[0] if email else ""
    return normalize_user_display_name(local_part.replace(".", " ").replace("_", " "))


def initialize_auth_state() -> None:
    defaults = {
        "autenticado": False,
        "usuario_email": "",
        "usuario_nombre": "",
        "es_admin": False,
        "rol_usuario": "captura",
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value


def current_user_role() -> str:
    return str(st.session_state.get("rol_usuario", "captura"))


def log_activity(accion: str, detalle: str = "", payload: dict[str, Any] | None = None) -> None:
    if not supabase_users_enabled():
        return

    form_key = ""
    equipment_code = ""
    month: int | None = None
    year: int | None = None
    target_payload = payload or st.session_state.get("payload")
    if isinstance(target_payload, dict):
        metadata = target_payload.get("metadata", {})
        form_key = str(metadata.get("form_key", "")).strip()
        equipment_code = str(metadata.get("equipment_code", "")).strip()
        try:
            month = int(metadata.get("month"))
            year = int(metadata.get("year"))
        except (TypeError, ValueError):
            month = None
            year = None

    try:
        registrar_evento_auditoria(
            email=str(st.session_state.get("usuario_email", "")).strip(),
            accion=accion,
            detalle=detalle,
            form_key=form_key,
            equipment_code=equipment_code,
            month=month,
            year=year,
        )
    except Exception:
        pass


def get_local_now() -> datetime:
    return datetime.now(LOCAL_TIMEZONE)


def format_local_timestamp(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    try:
        parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
        if parsed.tzinfo is None:
            parsed = parsed.replace(tzinfo=LOCAL_TIMEZONE)
        return parsed.astimezone(LOCAL_TIMEZONE).strftime("%Y-%m-%d %H:%M:%S")
    except ValueError:
        return text


def can_edit_sensitive_configuration() -> bool:
    return current_user_role() in SENSITIVE_EDITOR_ROLES


def can_edit_correction_settings() -> bool:
    return current_user_role() == "admin"


def can_edit_schedule() -> bool:
    return current_user_role() == "admin"


def can_edit_daily_records() -> bool:
    return current_user_role() in {"captura", "admin"}


def can_verify_daily_records() -> bool:
    return current_user_role() in {"captura", "admin"}


def can_close_period() -> bool:
    return current_user_role() in {"calidad", "admin"}


def can_manage_traceability() -> bool:
    return current_user_role() in {"calidad", "admin"}


def can_edit_master_list() -> bool:
    return current_user_role() in {"calidad", "admin"}


def can_export_period() -> bool:
    return current_user_role() in {"calidad", "admin"}


def is_capture_role() -> bool:
    return current_user_role() == "captura"


def is_admin_role() -> bool:
    return current_user_role() == "admin"


def render_auth_screen() -> None:
    st.title("Acceso al sistema")

    if not supabase_users_enabled():
        st.error("La autenticacion no esta configurada en esta app.")
        st.stop()

    login_tab, register_tab = st.tabs(["Iniciar sesion", "Crear cuenta"])

    with login_tab:
        email_login = st.text_input("Correo", key="login_email")
        password_login = st.text_input("Contrasena", type="password", key="login_password")
        if st.button("Ingresar", use_container_width=True):
            try:
                result = autenticar_usuario(email_login, password_login, normalize_user_role)
                if result["ok"]:
                    st.session_state["autenticado"] = True
                    st.session_state["usuario_email"] = result["email"]
                    st.session_state["usuario_nombre"] = normalize_user_display_name(
                        result.get("nombre") or get_known_display_name(result["email"])
                    )
                    st.session_state["es_admin"] = result["es_admin"]
                    st.session_state["rol_usuario"] = result.get("rol", "captura")
                    log_activity("inicio_sesion", "Ingreso a la app")
                    st.rerun()
                else:
                    st.error(result["mensaje"])
            except Exception as exc:
                st.error(f"No se pudo iniciar sesion: {exc}")

    with register_tab:
        email_register = st.text_input("Correo institucional o personal", key="register_email")
        name_register = st.text_input("Nombre para firma automatica", key="register_name")
        password_register = st.text_input("Contrasena", type="password", key="register_password")
        password_register_2 = st.text_input("Confirmar contrasena", type="password", key="register_password_2")
        requested_role = st.selectbox(
            "Perfil solicitado",
            ["captura", "calidad"],
            format_func=lambda value: value.capitalize(),
            key="register_role",
        )
        if st.button("Crear cuenta", use_container_width=True):
            try:
                if not email_register or not password_register:
                    st.warning("Completa correo y contrasena.")
                elif not name_register.strip():
                    st.warning("Captura el nombre para la firma automatica.")
                elif password_register != password_register_2:
                    st.warning("Las contrasenas no coinciden.")
                else:
                    registrar_usuario(
                        email_register,
                        password_register,
                        requested_role,
                        normalize_user_display_name(name_register),
                    )
                    st.success("Cuenta creada. Queda pendiente de aprobacion.")
            except Exception as exc:
                st.error(f"No se pudo crear la cuenta: {exc}")


def render_user_admin_sidebar() -> None:
    if not st.session_state.get("es_admin", False) or not supabase_users_enabled():
        return

    st.sidebar.divider()
    st.sidebar.subheader("Administracion")
    try:
        pending_users = obtener_usuarios_pendientes()
        st.sidebar.markdown("**Solicitudes pendientes**")
        if pending_users:
            for user_id, email, name, registered_at in pending_users:
                st.sidebar.write(f"{email} - {registered_at}")
                st.sidebar.caption(f"Nombre: {name or get_known_display_name(email) or 'Sin capturar'}")
                approval_role = st.sidebar.selectbox(
                    f"Rol para {email}",
                    ROLES_USUARIO,
                    index=ROLES_USUARIO.index("captura"),
                    format_func=lambda value: value.capitalize(),
                    key=f"approval_role_{user_id}",
                )
                if st.sidebar.button("Aprobar", key=f"approve_{user_id}", use_container_width=True):
                    aprobar_usuario(user_id, approval_role)
                    log_activity("aprobar_usuario", f"{email} -> {approval_role}")
                    st.sidebar.success(f"Usuario {email} aprobado.")
                    st.rerun()
        else:
            st.sidebar.caption("No hay usuarios pendientes.")

        st.sidebar.markdown("**Roles activos**")
        approved_users = [row for row in listar_usuarios() if row[3] == 1]
        admin_count = sum(1 for row in approved_users if row[4] == 1)
        if approved_users:
            for user_id, email, name, _, is_admin, role, _ in approved_users:
                normalized_role = normalize_user_role(role, bool(is_admin))
                display_name_key = f"name_user_{user_id}"
                if display_name_key not in st.session_state:
                    st.session_state[display_name_key] = normalize_user_display_name(name or get_known_display_name(email))
                edited_name = st.sidebar.text_input(
                    f"Nombre de {email}",
                    key=display_name_key,
                )
                new_role = st.sidebar.selectbox(
                    email,
                    ROLES_USUARIO,
                    index=ROLES_USUARIO.index(normalized_role),
                    format_func=lambda value: value.capitalize(),
                    key=f"role_user_{user_id}",
                )
                if st.sidebar.button("Actualizar usuario", key=f"update_role_{user_id}", use_container_width=True):
                    normalized_name = normalize_user_display_name(edited_name)
                    if normalized_name:
                        try:
                            actualizar_nombre_usuario(user_id, normalized_name)
                        except Exception:
                            pass
                    actualizar_rol_usuario(user_id, new_role)
                    log_activity("actualizar_usuario", f"{email} -> {new_role} | {normalized_name}")
                    st.sidebar.success(f"Usuario {email} actualizado.")
                    st.rerun()
                can_delete_user = email != str(st.session_state.get("usuario_email", "")).strip().lower()
                would_remove_last_admin = normalized_role == "admin" and admin_count <= 1
                if st.sidebar.button("Quitar acceso", key=f"delete_user_{user_id}", use_container_width=True):
                    if not can_delete_user:
                        st.sidebar.warning("No puedes quitar tu propio acceso desde aqui.")
                    elif would_remove_last_admin:
                        st.sidebar.warning("No puedes quitar al ultimo admin activo.")
                    else:
                        eliminar_usuario(user_id)
                        log_activity("quitar_acceso", email)
                        st.sidebar.success(f"Se quito el acceso de {email}.")
                        st.rerun()
        else:
            st.sidebar.caption("No hay usuarios aprobados para administrar.")

        with st.sidebar.expander("Historial de actividad", expanded=False):
            try:
                eventos = listar_eventos_auditoria(limit=40)
                if eventos:
                    for evento in eventos:
                        marca_tiempo = format_local_timestamp(str(evento.get("created_at", "")))
                        accion = str(evento.get("accion", "")).replace("_", " ").capitalize()
                        email = str(evento.get("email", ""))
                        detalle = str(evento.get("detalle", "")).strip()
                        contexto = []
                        if evento.get("form_key"):
                            contexto.append(str(evento["form_key"]))
                        if evento.get("equipment_code"):
                            contexto.append(str(evento["equipment_code"]))
                        if evento.get("year") and evento.get("month"):
                            contexto.append(f"{int(evento['year'])}-{int(evento['month']):02d}")
                        context_label = " | ".join(contexto)
                        st.write(f"{marca_tiempo} - {email}")
                        st.caption(f"{accion}{' | ' + detalle if detalle else ''}{' | ' + context_label if context_label else ''}")
                else:
                    st.caption("Sin actividad registrada todavia.")
            except Exception:
                st.caption("El historial aun no esta disponible.")
    except Exception as exc:
        st.sidebar.error(f"No se pudo cargar la administracion de usuarios: {exc}")


def coerce_factor_value(raw_value: Any) -> tuple[str, float]:
    try:
        numeric_value = float(str(raw_value).replace(",", ".").strip())
    except (TypeError, ValueError, AttributeError):
        numeric_value = 0.0

    operation = "+" if numeric_value >= 0 else "-"
    return operation, abs(numeric_value)


def parse_range_bounds(label_text: str) -> tuple[float, float] | None:
    normalized_text = str(label_text).strip().lower()
    match = re.search(
        r"(-?\d+(?:\.\d+)?)\s*(?:-|a|to|–|—)\s*(-?\d+(?:\.\d+)?)",
        normalized_text,
    )
    if not match:
        return None
    return float(match.group(1)), float(match.group(2))


def infer_range_separator(label_text: str) -> str:
    normalized_text = str(label_text).strip().lower()
    if " a " in normalized_text:
        return "a"
    if " to " in normalized_text:
        return "to"
    return "-"


def format_range_number(value: float) -> str:
    return format_decimal_value(float(value))


def build_range_label(min_value: float, max_value: float, separator: str = "-") -> str:
    left = format_range_number(min_value)
    right = format_range_number(max_value)
    if separator == "a":
        return f"{left} a {right}"
    if separator == "to":
        return f"{left} to {right}"
    return f"{left} - {right}"


def find_value_after_label(
    worksheet: Any,
    label: str,
    rows: list[int],
    max_columns: int | None = None,
) -> Any:
    max_col = max_columns or worksheet.max_column
    normalized_label = label.strip().lower()

    for row in rows:
        for column in range(1, max_col + 1):
            cell_value = worksheet.cell(row, column).value
            if cell_value is None:
                continue
            if str(cell_value).strip().lower() != normalized_label:
                continue

            for next_col in range(column + 1, max_col + 1):
                next_value = worksheet.cell(row, next_col).value
                if next_value not in (None, ""):
                    return next_value

    return None


def find_cell_after_label(
    worksheet: Any,
    label: str,
    rows: list[int],
    max_columns: int | None = None,
) -> str | None:
    max_col = max_columns or worksheet.max_column
    normalized_label = label.strip().lower()

    for row in rows:
        for column in range(1, max_col + 1):
            cell_value = worksheet.cell(row, column).value
            if cell_value is None:
                continue
            if str(cell_value).strip().lower() != normalized_label:
                continue

            for next_col in range(column + 1, max_col + 1):
                next_value = worksheet.cell(row, next_col).value
                if next_value not in (None, ""):
                    return worksheet.cell(row, next_col).coordinate

    return None


def get_form_definition(form_key: str) -> dict[str, Any]:
    return FORM_DEFINITIONS[form_key]


def get_template_paths(form_key: str) -> tuple[Path, Path]:
    definition = get_form_definition(form_key)
    return BASE_DIR / definition["source_file"], BASE_DIR / definition["working_file"]


def _read_local_template_bytes(source_path: Path, working_path: Path) -> bytes:
    try:
        return source_path.read_bytes()
    except PermissionError:
        if not working_path.exists() and source_path.exists():
            try:
                shutil.copy2(source_path, working_path)
            except PermissionError:
                workbook = load_workbook(source_path)
                workbook.save(working_path)
        return working_path.read_bytes()
    except FileNotFoundError:
        if working_path.exists():
            return working_path.read_bytes()
        raise


@st.cache_data(show_spinner=False, ttl=300)
def _get_template_bytes_cached(
    form_key: str,
    source_file_name: str,
    _templates_storage_cache_key: tuple[bool, str, str],
) -> bytes:
    source_path, working_path = get_template_paths(form_key)
    if templates_storage_enabled():
        try:
            return download_template_bytes(source_file_name)
        except Exception:
            pass
    return _read_local_template_bytes(source_path, working_path)


def get_template_bytes(form_key: str) -> bytes:
    definition = get_form_definition(form_key)
    return _get_template_bytes_cached(
        form_key,
        str(definition["source_file"]),
        get_templates_storage_cache_key(),
    )


def template_source_available(form_key: str) -> bool:
    source_path, working_path = get_template_paths(form_key)
    if templates_storage_enabled():
        return True
    return source_path.exists() or working_path.exists()


def extract_inline_corrections(
    worksheet: Any,
    range_row: int,
    factor_row: int,
) -> tuple[
    dict[str, dict[str, Any]],
    dict[str, str],
    dict[str, str],
    dict[str, str],
    dict[str, float],
]:
    correction_bands: dict[str, dict[str, Any]] = {}
    correction_range_cells: dict[str, str] = {}
    correction_cells: dict[str, str] = {}
    correction_operations: dict[str, str] = {}
    correction_factors: dict[str, float] = {}
    range_index = 1
    range_header_col = None

    for column in range(1, worksheet.max_column + 1):
        header_value = worksheet.cell(range_row, column).value
        if header_value and "Rango" in str(header_value):
            range_header_col = column
            break

    if range_header_col is None:
        return correction_bands, correction_range_cells, correction_cells, correction_operations, correction_factors

    for column in range(range_header_col + 1, worksheet.max_column + 1):
        range_value = worksheet.cell(range_row, column).value
        factor_value = worksheet.cell(factor_row, column).value
        if range_value in (None, "") or factor_value in (None, ""):
            continue

        label_text = str(range_value).strip()
        if label_text.upper() == "N/A":
            continue
        if "-" not in label_text:
            continue

        bounds = parse_range_bounds(label_text)
        if bounds is None:
            continue
        min_value, max_value = bounds

        key = f"range_{range_index}"
        range_index += 1
        operation, numeric_value = coerce_factor_value(factor_value)
        correction_bands[key] = {
            "label": label_text,
            "min": min_value,
            "max": max_value,
            "separator": infer_range_separator(label_text),
        }
        correction_range_cells[key] = worksheet.cell(range_row, column).coordinate
        correction_cells[key] = worksheet.cell(factor_row, column).coordinate
        correction_operations[key] = operation
        correction_factors[key] = numeric_value

    return correction_bands, correction_range_cells, correction_cells, correction_operations, correction_factors


def extract_cold_equipment_config(sheet_name: str, worksheet: Any) -> dict[str, Any]:
    (
        correction_bands,
        correction_range_cells,
        correction_cells,
        correction_operations,
        correction_factors,
    ) = extract_inline_corrections(
        worksheet,
        range_row=6,
        factor_row=7,
    )
    return {
        "sheet_name": sheet_name,
        "equipment_code": sheet_name,
        "laboratory": find_value_after_label(worksheet, "Laboratorio", [6]) or "",
        "equipment_name": find_value_after_label(worksheet, "Equipo", [7]) or "",
        "brand": find_value_after_label(worksheet, "Marca", [7]) or "",
        "model": find_value_after_label(worksheet, "Modelo", [7]) or "",
        "serial_number": find_value_after_label(worksheet, "No. Serie", [6]) or "",
        "inventory_code": find_value_after_label(worksheet, "Inventario / Código", [7])
        or find_value_after_label(worksheet, "Inventario/Código", [7])
        or "",
        "primary_label": find_value_after_label(worksheet, "Temperatura (ºC)", [8]) or "",
        "minimum_label": find_value_after_label(worksheet, "Mínima", [8]) or "",
        "maximum_label": find_value_after_label(worksheet, "Máxima", [8]) or "",
        "metadata_cells": {
            "equipment_name": find_cell_after_label(worksheet, "Equipo", [7]),
            "brand": find_cell_after_label(worksheet, "Marca", [7]),
            "model": find_cell_after_label(worksheet, "Modelo", [7]),
            "serial_number": find_cell_after_label(worksheet, "No. Serie", [6]),
            "inventory_code": find_cell_after_label(worksheet, "Inventario / Código", [7])
            or find_cell_after_label(worksheet, "Inventario/Código", [7]),
            "temperature_label": find_cell_after_label(worksheet, "Temperatura (ºC)", [8]),
            "minimum_label": find_cell_after_label(worksheet, "Mínima", [8]),
            "maximum_label": find_cell_after_label(worksheet, "Máxima", [8]),
        },
        "correction_bands": correction_bands,
        "correction_range_cells": correction_range_cells,
        "correction_cells": correction_cells,
        "correction_factors": correction_factors,
        "correction_operations": correction_operations,
    }


def extract_incubator_config(sheet_name: str, worksheet: Any) -> dict[str, Any]:
    (
        correction_bands,
        correction_range_cells,
        correction_cells,
        correction_operations,
        correction_factors,
    ) = extract_inline_corrections(
        worksheet,
        range_row=6,
        factor_row=7,
    )
    return {
        "sheet_name": sheet_name,
        "equipment_code": sheet_name,
        "laboratory": find_value_after_label(worksheet, "Laboratorio", [6]) or "",
        "equipment_name": find_value_after_label(worksheet, "Equipo", [7]) or "",
        "brand": find_value_after_label(worksheet, "Marca", [7]) or "",
        "model": find_value_after_label(worksheet, "Modelo", [7]) or "",
        "serial_number": find_value_after_label(worksheet, "No. Serie", [6]) or "",
        "inventory_code": find_value_after_label(worksheet, "Inventario / Código", [7])
        or find_value_after_label(worksheet, "Inventario/Código", [7])
        or "",
        "primary_label": find_value_after_label(worksheet, "Temperatura (ºC)", [8]) or "",
        "minimum_label": find_value_after_label(worksheet, "Mínima", [8]) or "",
        "maximum_label": find_value_after_label(worksheet, "Máxima", [8]) or "",
        "secondary_label": find_value_after_label(worksheet, "%CO2", [9]) or "",
        "secondary_minimum_label": find_value_after_label(worksheet, "Mínima", [9]) or "",
        "secondary_maximum_label": find_value_after_label(worksheet, "Máxima", [9]) or "",
        "metadata_cells": {
            "equipment_name": find_cell_after_label(worksheet, "Equipo", [7]),
            "brand": find_cell_after_label(worksheet, "Marca", [7]),
            "model": find_cell_after_label(worksheet, "Modelo", [7]),
            "serial_number": find_cell_after_label(worksheet, "No. Serie", [6]),
            "inventory_code": find_cell_after_label(worksheet, "Inventario / Código", [7])
            or find_cell_after_label(worksheet, "Inventario/Código", [7]),
            "temperature_label": find_cell_after_label(worksheet, "Temperatura (ºC)", [8]),
            "minimum_label": find_cell_after_label(worksheet, "Mínima", [8]),
            "maximum_label": find_cell_after_label(worksheet, "Máxima", [8]),
            "secondary_label": find_cell_after_label(worksheet, "%CO2", [9]),
            "secondary_minimum_label": find_cell_after_label(worksheet, "Mínima", [9]),
            "secondary_maximum_label": find_cell_after_label(worksheet, "Máxima", [9]),
        },
        "correction_bands": correction_bands,
        "correction_range_cells": correction_range_cells,
        "correction_cells": correction_cells,
        "correction_factors": correction_factors,
        "correction_operations": correction_operations,
    }


def extract_ambient_config(sheet_name: str, worksheet: Any) -> dict[str, Any]:
    is_humidity_sheet = "HUMEDAD" in sheet_name.upper()
    if is_humidity_sheet:
        uses_shifted_header = worksheet["O8"].value in (None, "") and worksheet["S8"].value in (None, "")
        if uses_shifted_header:
            brand_value = worksheet["Q8"].value or ""
            model_value = worksheet["W8"].value or ""
            serial_value = worksheet["AC8"].value or ""
            brand_cell = "Q8"
            model_cell = "W8"
            serial_cell = "AC8"
        else:
            brand_value = worksheet["O8"].value or ""
            model_value = worksheet["S8"].value or ""
            serial_value = worksheet["W8"].value or ""
            brand_cell = "O8"
            model_cell = "S8"
            serial_cell = "W8"
    else:
        brand_value = worksheet["Q8"].value or ""
        model_value = worksheet["W8"].value or ""
        serial_value = worksheet["AC8"].value or ""
        brand_cell = "Q8"
        model_cell = "W8"
        serial_cell = "AC8"
    primary_label_value = (
        find_value_after_label(worksheet, "TEMPERATURA (°C)", [9])
        or find_value_after_label(worksheet, "TEMPERATURA (Â°C)", [9])
        or find_value_after_label(worksheet, "% HUMEDAD", [9])
        or find_value_after_label(worksheet, "% Humedad Relativa", [9])
        or ""
    )
    primary_label_cell = (
        find_cell_after_label(worksheet, "TEMPERATURA (°C)", [9])
        or find_cell_after_label(worksheet, "TEMPERATURA (Â°C)", [9])
        or find_cell_after_label(worksheet, "% HUMEDAD", [9])
        or find_cell_after_label(worksheet, "% Humedad Relativa", [9])
    )
    correction_bands: dict[str, dict[str, Any]] = {}
    correction_range_cells: dict[str, str] = {}
    correction_cells: dict[str, str] = {}
    operations: dict[str, str] = {}
    factors: dict[str, float] = {}
    range_index = 1
    for column in range(1, worksheet.max_column + 1):
        range_label = worksheet.cell(7, column).value
        factor_value = worksheet.cell(8, column).value
        if not range_label or factor_value in (None, ""):
            continue
        label_text = str(range_label).strip()
        if "-" not in label_text:
            continue
        bounds = parse_range_bounds(label_text)
        if bounds is None:
            continue
        min_value, max_value = bounds

        key = f"range_{range_index}"
        range_index += 1
        op, num = coerce_factor_value(factor_value)
        operations[key] = op
        factors[key] = num
        correction_bands[key] = {
            "label": label_text,
            "min": min_value,
            "max": max_value,
            "separator": infer_range_separator(label_text),
        }
        correction_range_cells[key] = worksheet.cell(7, column).coordinate
        correction_cells[key] = worksheet.cell(8, column).coordinate
    return {
        "sheet_name": sheet_name,
        "equipment_code": sheet_name,
        "laboratory": find_value_after_label(worksheet, "Laboratorio", [6]) or "",
        "equipment_name": worksheet["E8"].value or sheet_name,
        "brand": brand_value,
        "model": model_value,
        "serial_number": serial_value,
        "inventory_code": worksheet["K8"].value or "",
        "primary_label": primary_label_value,
        "minimum_label": find_value_after_label(worksheet, "Mínima", [9]) or "",
        "maximum_label": find_value_after_label(worksheet, "Máxima", [9]) or "",
        "metadata_cells": {
            "equipment_name": "E8",
            "brand": brand_cell,
            "model": model_cell,
            "serial_number": serial_cell,
            "inventory_code": "K8",
            "temperature_label": primary_label_cell,
            "minimum_label": find_cell_after_label(worksheet, "Mínima", [9]),
            "maximum_label": find_cell_after_label(worksheet, "Máxima", [9]),
        },
        "correction_bands": correction_bands,
        "correction_range_cells": correction_range_cells,
        "correction_cells": correction_cells,
        "correction_factors": factors,
        "correction_operations": operations,
    }


@st.cache_data(show_spinner=False)
def _load_equipment_configs_cached(
    form_key: str,
    templates_cache_key: str,
    cache_version: str,
) -> dict[str, dict[str, Any]]:
    _ = templates_cache_key, cache_version
    workbook = load_template_workbook(form_key, data_only=True)
    equipment_configs: dict[str, dict[str, Any]] = {}
    definition = get_form_definition(form_key)

    for sheet_name in workbook.sheetnames:
        if sheet_name in definition["sheet_exclusions"]:
            continue

        worksheet = workbook[sheet_name]
        extractor = definition["extractor"]
        if extractor == "cold_equipment":
            equipment_configs[sheet_name] = extract_cold_equipment_config(sheet_name, worksheet)
        elif extractor == "incubators":
            equipment_configs[sheet_name] = extract_incubator_config(sheet_name, worksheet)
        else:
            equipment_configs[sheet_name] = extract_ambient_config(sheet_name, worksheet)

        if not definition["supports_corrections"]:
            equipment_configs[sheet_name]["correction_bands"] = {}
            equipment_configs[sheet_name]["correction_range_cells"] = {}
            equipment_configs[sheet_name]["correction_cells"] = {}
            equipment_configs[sheet_name]["correction_factors"] = {}
            equipment_configs[sheet_name]["correction_operations"] = {}

    return equipment_configs


def load_equipment_configs(form_key: str) -> dict[str, dict[str, Any]]:
    return _load_equipment_configs_cached(
        form_key,
        get_templates_storage_cache_key(),
        EQUIPMENT_CONFIG_CACHE_VERSION,
    )


def build_default_payload(
    equipment_code: str = DEFAULT_EQUIPMENT_CODE,
    form_key: str = DEFAULT_FORM_KEY,
) -> dict[str, Any]:
    today = date.today()
    equipment_configs = load_equipment_configs(form_key)
    equipment_config = equipment_configs[equipment_code]
    definition = get_form_definition(form_key)
    return {
        "metadata": {
            "form_key": form_key,
            "form_label": definition["label"],
            "month": today.month,
            "year": today.year,
            "equipment_code": equipment_config["equipment_code"],
            "laboratory": equipment_config["laboratory"],
            "equipment_name": equipment_config["equipment_name"],
            "brand": equipment_config["brand"],
            "model": equipment_config["model"],
            "serial_number": equipment_config["serial_number"],
            "inventory_code": equipment_config["inventory_code"],
            "temperature_label": equipment_config["primary_label"],
            "minimum_label": equipment_config["minimum_label"],
            "maximum_label": equipment_config["maximum_label"],
            "secondary_label": equipment_config.get("secondary_label", ""),
            "secondary_minimum_label": equipment_config.get("secondary_minimum_label", ""),
            "secondary_maximum_label": equipment_config.get("secondary_maximum_label", ""),
        },
        "metadata_cells": dict(equipment_config.get("metadata_cells", {})),
        "correction_bands": dict(equipment_config.get("correction_bands", {})),
        "correction_range_cells": dict(equipment_config.get("correction_range_cells", {})),
        "correction_cells": dict(equipment_config.get("correction_cells", {})),
        "correction_factors": dict(equipment_config["correction_factors"]),
        "correction_operations": dict(equipment_config["correction_operations"]),
        "non_working_days": [],
        "daily_records": {
            str(day): asdict(default_daily_capture(day)) for day in range(1, 32)
        },
        "change_log": [],
        "monthly_closure": {
            "observations": "",
            "reviewed_by": "",
            "reviewed_on": today.isoformat(),
        },
    }


def hydrate_payload_corrections(payload: dict[str, Any]) -> dict[str, Any]:
    form_key = str(payload.get("metadata", {}).get("form_key", DEFAULT_FORM_KEY))
    equipment_code = str(payload.get("metadata", {}).get("equipment_code", DEFAULT_EQUIPMENT_CODE))
    equipment_config = load_equipment_configs(form_key).get(equipment_code)
    if not equipment_config:
        return payload

    template_bands = dict(equipment_config.get("correction_bands", {}))
    template_range_cells = dict(equipment_config.get("correction_range_cells", {}))
    template_cells = dict(equipment_config.get("correction_cells", {}))
    template_factors = dict(equipment_config.get("correction_factors", {}))
    template_operations = dict(equipment_config.get("correction_operations", {}))

    current_bands = dict(payload.get("correction_bands", {}))
    hydrated_bands: dict[str, dict[str, Any]] = {}
    for factor_key, template_band in template_bands.items():
        current_band = current_bands.get(factor_key, {})
        min_value = float(current_band.get("min", template_band.get("min", 0)))
        max_value = float(current_band.get("max", template_band.get("max", 0)))
        separator = str(current_band.get("separator", template_band.get("separator", infer_range_separator(str(template_band.get("label", "")))))).strip() or "-"
        hydrated_bands[factor_key] = {
            "label": build_range_label(min_value, max_value, separator),
            "min": min_value,
            "max": max_value,
            "separator": separator,
        }

    payload["correction_bands"] = hydrated_bands
    payload["correction_range_cells"] = template_range_cells
    payload["correction_cells"] = template_cells

    current_factors = dict(payload.get("correction_factors", {}))
    current_operations = dict(payload.get("correction_operations", {}))

    hydrated_factors: dict[str, Any] = {}
    hydrated_operations: dict[str, str] = {}
    for factor_key, factor_value in template_factors.items():
        hydrated_factors[factor_key] = current_factors.get(factor_key, factor_value)
        hydrated_operations[factor_key] = str(current_operations.get(factor_key, template_operations.get(factor_key, "+")))

    payload["correction_factors"] = hydrated_factors
    payload["correction_operations"] = hydrated_operations
    return payload


def ensure_data_dir() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    TRACEABILITY_DIR.mkdir(parents=True, exist_ok=True)
    EQUIPMENT_CONFIG_DIR.mkdir(parents=True, exist_ok=True)


def extract_correction_settings(payload: dict[str, Any]) -> dict[str, Any]:
    return {
        "correction_bands": dict(payload.get("correction_bands", {})),
        "correction_factors": dict(payload.get("correction_factors", {})),
        "correction_operations": dict(payload.get("correction_operations", {})),
    }


def correction_settings_are_empty(settings: dict[str, Any] | None) -> bool:
    if not isinstance(settings, dict):
        return True
    return not any(
        settings.get(key)
        for key in ("correction_bands", "correction_factors", "correction_operations")
    )


def get_equipment_config_file(form_key: str, equipment_code: str) -> Path:
    ensure_data_dir()
    safe_form = re.sub(r"[^a-z0-9_]+", "_", form_key.lower())
    safe_equipment = re.sub(r"[^a-z0-9_]+", "_", equipment_code.lower())
    return EQUIPMENT_CONFIG_DIR / f"{safe_form}_{safe_equipment}_corrections.json"


def load_local_correction_settings(form_key: str, equipment_code: str) -> dict[str, Any] | None:
    config_file = get_equipment_config_file(form_key, equipment_code)
    if not config_file.exists():
        return None
    try:
        data = json.loads(config_file.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None
    return data if isinstance(data, dict) else None


def save_local_correction_settings(form_key: str, equipment_code: str, settings: dict[str, Any]) -> None:
    get_equipment_config_file(form_key, equipment_code).write_text(
        json.dumps(settings, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def get_global_period_config_file(year: int, month: int) -> Path:
    ensure_data_dir()
    return EQUIPMENT_CONFIG_DIR / f"global_non_working_days_{int(year)}_{int(month):02d}.json"


def normalize_non_working_days(value: Any) -> list[int]:
    if not isinstance(value, list):
        return []
    days: set[int] = set()
    for item in value:
        try:
            day = int(item)
        except (TypeError, ValueError):
            continue
        if 1 <= day <= 31:
            days.add(day)
    return sorted(days)


def load_shared_non_working_days(year: int, month: int) -> list[int] | None:
    config_type = f"non_working_days:{int(year)}:{int(month):02d}"
    if supabase_storage_enabled():
        try:
            payload = load_remote_equipment_config_payload("__global__", "__global__", config_type)
            if isinstance(payload, dict):
                return normalize_non_working_days(payload.get("non_working_days", []))
        except Exception:
            pass

    config_file = get_global_period_config_file(year, month)
    if config_file.exists():
        try:
            payload = json.loads(config_file.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            payload = {}
        if isinstance(payload, dict):
            return normalize_non_working_days(payload.get("non_working_days", []))
    return None


def save_shared_non_working_days(year: int, month: int, non_working_days: list[int]) -> str:
    normalized_days = normalize_non_working_days(non_working_days)
    payload = {"year": int(year), "month": int(month), "non_working_days": normalized_days}
    config_type = f"non_working_days:{int(year)}:{int(month):02d}"
    if supabase_storage_enabled():
        try:
            save_remote_equipment_config_payload(
                "__global__",
                "__global__",
                payload,
                config_type=config_type,
                updated_by=str(st.session_state.get("usuario_email", "")).strip(),
            )
            return "Supabase"
        except Exception:
            pass

    get_global_period_config_file(year, month).write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return "Local JSON"


def load_shared_correction_settings(form_key: str, equipment_code: str) -> dict[str, Any] | None:
    ensure_data_dir()
    if supabase_storage_enabled():
        try:
            settings = load_remote_equipment_config_payload(form_key, equipment_code, "corrections")
            if not correction_settings_are_empty(settings):
                return settings
        except Exception:
            pass

    settings = load_local_correction_settings(form_key, equipment_code)
    if not correction_settings_are_empty(settings):
        return settings
    return None


def find_latest_saved_correction_settings(form_key: str, equipment_code: str) -> dict[str, Any] | None:
    periods = sorted(
        list_periods_for_equipment(form_key, equipment_code),
        key=lambda row: (int(row.get("year", 0)), int(row.get("month", 0))),
        reverse=True,
    )
    for period in periods:
        try:
            year = int(period["year"])
            month = int(period["month"])
        except (KeyError, TypeError, ValueError):
            continue

        data: dict[str, Any] | None = None
        if supabase_storage_enabled():
            try:
                data = load_remote_period_payload(form_key, equipment_code, year, month)
            except Exception:
                data = None
        if data is None:
            data_file = get_period_file_from_values(form_key, equipment_code, year, month)
            if data_file.exists():
                try:
                    data = json.loads(data_file.read_text(encoding="utf-8"))
                except (OSError, json.JSONDecodeError):
                    data = None
        if not isinstance(data, dict):
            continue

        settings = extract_correction_settings(data)
        if not correction_settings_are_empty(settings):
            return settings
    return None


def save_shared_correction_settings(payload: dict[str, Any]) -> str:
    metadata = payload.get("metadata", {})
    form_key = str(metadata.get("form_key", "")).strip()
    equipment_code = str(metadata.get("equipment_code", "")).strip()
    if not form_key or not equipment_code:
        return "omitido"

    definition = get_form_definition(form_key)
    if not definition.get("supports_corrections"):
        return "omitido"

    settings = extract_correction_settings(payload)
    if correction_settings_are_empty(settings):
        return "omitido"

    if supabase_storage_enabled():
        try:
            save_remote_equipment_config_payload(
                form_key,
                equipment_code,
                settings,
                config_type="corrections",
                updated_by=str(st.session_state.get("usuario_email", "")).strip(),
            )
            return "Supabase"
        except Exception:
            pass

    save_local_correction_settings(form_key, equipment_code, settings)
    return "Local JSON"


def apply_shared_correction_settings(payload: dict[str, Any]) -> dict[str, Any]:
    metadata = payload.get("metadata", {})
    form_key = str(metadata.get("form_key", DEFAULT_FORM_KEY))
    equipment_code = str(metadata.get("equipment_code", DEFAULT_EQUIPMENT_CODE))
    shared_settings = load_shared_correction_settings(form_key, equipment_code)
    if correction_settings_are_empty(shared_settings):
        shared_settings = find_latest_saved_correction_settings(form_key, equipment_code)
        if not correction_settings_are_empty(shared_settings):
            payload_for_save = dict(payload)
            payload_for_save["correction_bands"] = shared_settings.get("correction_bands", {})
            payload_for_save["correction_factors"] = shared_settings.get("correction_factors", {})
            payload_for_save["correction_operations"] = shared_settings.get("correction_operations", {})
            save_shared_correction_settings(payload_for_save)
    if correction_settings_are_empty(shared_settings):
        return payload

    payload["correction_bands"] = shared_settings.get("correction_bands", payload.get("correction_bands", {}))
    payload["correction_factors"].update(shared_settings.get("correction_factors", {}))
    payload["correction_operations"].update(shared_settings.get("correction_operations", {}))
    return hydrate_payload_corrections(payload)


def get_traceability_file(form_key: str, equipment_code: str) -> Path:
    ensure_data_dir()
    safe_form = re.sub(r"[^a-z0-9_]+", "_", form_key.lower())
    safe_equipment = re.sub(r"[^a-z0-9_]+", "_", equipment_code.lower())
    return TRACEABILITY_DIR / f"{safe_form}_{safe_equipment}.json"


def normalize_traceability_entry(entry: dict[str, Any]) -> dict[str, Any]:
    scheduled_for = str(entry.get("scheduled_for", "")).strip()
    completed_on = str(entry.get("completed_on", "")).strip()
    return {
        "id": str(entry.get("id", "")).strip() or str(uuid4()),
        "form_key": str(entry.get("form_key", "")).strip(),
        "equipment_code": str(entry.get("equipment_code", "")).strip(),
        "entry_type": str(entry.get("entry_type", "")).strip(),
        "status": str(entry.get("status", "")).strip() or "programado",
        "scheduled_for": scheduled_for,
        "completed_on": completed_on,
        "provider": str(entry.get("provider", "")).strip(),
        "notes": str(entry.get("notes", "")).strip(),
        "created_by": str(entry.get("created_by", "")).strip(),
        "updated_by": str(entry.get("updated_by", "")).strip(),
        "updated_at": str(entry.get("updated_at", "")).strip(),
    }


def list_traceability_entries(form_key: str, equipment_code: str) -> list[dict[str, Any]]:
    ensure_data_dir()
    if supabase_storage_enabled():
        try:
            return list_remote_traceability_entries(form_key, equipment_code)
        except Exception:
            pass

    traceability_file = get_traceability_file(form_key, equipment_code)
    if not traceability_file.exists():
        return []
    try:
        data = json.loads(traceability_file.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return []
    if not isinstance(data, list):
        return []
    entries = [normalize_traceability_entry(entry) for entry in data if isinstance(entry, dict)]
    entries.sort(key=lambda item: (item.get("scheduled_for") or "9999-12-31", item.get("entry_type") or ""))
    return entries


def save_traceability_entry(entry: dict[str, Any]) -> dict[str, Any]:
    ensure_data_dir()
    normalized_entry = normalize_traceability_entry(entry)
    normalized_entry["updated_by"] = str(st.session_state.get("usuario_email", "")).strip().lower()
    normalized_entry["updated_at"] = datetime.now(LOCAL_TIMEZONE).isoformat()
    if not normalized_entry["created_by"]:
        normalized_entry["created_by"] = normalized_entry["updated_by"]
    if supabase_storage_enabled():
        try:
            return save_remote_traceability_entry(normalized_entry, updated_by=normalized_entry["updated_by"])
        except Exception:
            pass

    entries = list_traceability_entries(normalized_entry["form_key"], normalized_entry["equipment_code"])
    updated_entries = [item for item in entries if item["id"] != normalized_entry["id"]]
    updated_entries.append(normalized_entry)
    updated_entries.sort(key=lambda item: (item.get("scheduled_for") or "9999-12-31", item.get("entry_type") or ""))
    get_traceability_file(normalized_entry["form_key"], normalized_entry["equipment_code"]).write_text(
        json.dumps(updated_entries, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return normalized_entry


def delete_traceability_entry(entry_id: str, form_key: str, equipment_code: str) -> None:
    normalized_id = entry_id.strip()
    if not normalized_id:
        return
    if supabase_storage_enabled():
        try:
            delete_remote_traceability_entry(normalized_id)
            return
        except Exception:
            pass

    entries = [item for item in list_traceability_entries(form_key, equipment_code) if item["id"] != normalized_id]
    get_traceability_file(form_key, equipment_code).write_text(
        json.dumps(entries, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def list_periods_for_equipment(form_key: str, equipment_code: str) -> list[dict[str, Any]]:
    ensure_data_dir()
    periods: list[dict[str, Any]] = []
    if supabase_storage_enabled():
        try:
            rows = list_remote_periods()
            periods = [
                row for row in rows
                if str(row.get("form_key", "")) == form_key and str(row.get("equipment_code", "")) == equipment_code
            ]
        except Exception:
            periods = []
    else:
        for data_file in DATA_DIR.glob("*.json"):
            try:
                data = json.loads(data_file.read_text(encoding="utf-8"))
            except (OSError, json.JSONDecodeError):
                continue
            metadata = data.get("metadata", {})
            if str(metadata.get("form_key", "")) != form_key or str(metadata.get("equipment_code", "")) != equipment_code:
                continue
            try:
                periods.append(
                    {
                        "form_key": form_key,
                        "equipment_code": equipment_code,
                        "month": int(metadata.get("month", 0)),
                        "year": int(metadata.get("year", 0)),
                    }
                )
            except (TypeError, ValueError):
                continue
    periods.sort(key=lambda row: (int(row["year"]), int(row["month"])))
    return periods


def normalize_canceled_slots(value: Any) -> list[int]:
    if value in (None, ""):
        return []
    if isinstance(value, str):
        raw_values = [item.strip() for item in value.split(",")]
    elif isinstance(value, list):
        raw_values = value
    else:
        raw_values = [value]

    canceled_slots: set[int] = set()
    for item in raw_values:
        try:
            index = int(item)
        except (TypeError, ValueError):
            continue
        if 0 <= index < len(TIME_SLOTS):
            canceled_slots.add(index)
    return sorted(canceled_slots)


def is_slot_canceled(record: dict[str, Any], slot_index: int) -> bool:
    return slot_index in normalize_canceled_slots(record.get("canceled_slots", []))


def is_day_fully_canceled(record: dict[str, Any]) -> bool:
    return len(normalize_canceled_slots(record.get("canceled_slots", []))) >= len(TIME_SLOTS)


def active_slot_indices(record: dict[str, Any]) -> list[int]:
    return [index for index in range(len(TIME_SLOTS)) if not is_slot_canceled(record, index)]


def merge_payload_with_saved_data(
    data: dict[str, Any],
    equipment_code: str = DEFAULT_EQUIPMENT_CODE,
    form_key: str = DEFAULT_FORM_KEY,
) -> dict[str, Any]:
    default_payload = build_default_payload(equipment_code=equipment_code, form_key=form_key)
    default_payload["metadata"].update(data.get("metadata", {}))
    default_payload["metadata_cells"] = data.get("metadata_cells", default_payload.get("metadata_cells", {}))
    default_payload["correction_bands"] = data.get("correction_bands", default_payload.get("correction_bands", {}))
    default_payload["correction_factors"].update(data.get("correction_factors", {}))
    default_payload["correction_operations"].update(data.get("correction_operations", {}))
    default_payload["non_working_days"] = data.get("non_working_days", [])
    default_payload["change_log"] = data.get("change_log", [])
    default_payload["monthly_closure"].update(data.get("monthly_closure", {}))

    incoming_records = data.get("daily_records", {})
    for day in range(1, 32):
        day_key = str(day)
        default_payload["daily_records"][day_key].update(incoming_records.get(day_key, {}))
        record = default_payload["daily_records"][day_key]
        if "temperatures" in record and "measured_temperatures" not in incoming_records.get(day_key, {}):
            record["measured_temperatures"] = record.pop("temperatures")
        record.setdefault("measured_temperatures", ["", "", ""])
        record.setdefault("corrected_temperatures", ["", "", ""])
        record.setdefault("secondary_measurements", ["", "", ""])
        record.setdefault("performed_by_slots", ["", "", ""])
        record["canceled_slots"] = normalize_canceled_slots(record.get("canceled_slots", []))
        record.setdefault("cancellation_note", "")
        if not str(record.get("recorded_on_mode", "")).strip():
            record["recorded_on_mode"] = "auto" if not str(record.get("verified_by", "")).strip() else "manual"

    return default_payload


def get_period_key(payload: dict[str, Any]) -> str:
    return (
        f"{payload['metadata']['form_key']}_"
        f"{payload['metadata']['equipment_code']}_"
        f"{int(payload['metadata']['year'])}_{int(payload['metadata']['month']):02d}"
    )


def get_period_file_from_values(form_key: str, equipment_code: str, year: int, month: int) -> Path:
    ensure_data_dir()
    normalized_form = form_key.lower().replace("-", "_")
    normalized_code = equipment_code.lower().replace("-", "_")
    return DATA_DIR / f"{normalized_form}_{normalized_code}_{year}_{month:02d}.json"


def get_period_file(payload: dict[str, Any]) -> Path:
    return get_period_file_from_values(
        str(payload["metadata"]["form_key"]),
        str(payload["metadata"]["equipment_code"]),
        int(payload["metadata"]["year"]),
        int(payload["metadata"]["month"]),
    )


def clear_period_widget_state(period_key: str) -> None:
    keys_to_delete = [
        key for key in st.session_state.keys() if period_key in str(key)
    ]
    for key in keys_to_delete:
        del st.session_state[key]


def load_saved_payload(
    form_key: str = DEFAULT_FORM_KEY,
    equipment_code: str = DEFAULT_EQUIPMENT_CODE,
    year: int | None = None,
    month: int | None = None,
) -> dict[str, Any]:
    default_payload = build_default_payload(equipment_code=equipment_code, form_key=form_key)
    if year is None:
        year = int(default_payload["metadata"]["year"])
    if month is None:
        month = int(default_payload["metadata"]["month"])

    if supabase_storage_enabled():
        try:
            remote_payload = load_remote_period_payload(form_key, equipment_code, year, month)
            if remote_payload:
                return apply_shared_correction_settings(
                    hydrate_payload_corrections(
                        merge_payload_with_saved_data(remote_payload, equipment_code=equipment_code, form_key=form_key)
                    )
                )
        except Exception:
            pass

    data_file = get_period_file_from_values(form_key, equipment_code, year, month)
    if not data_file.exists():
        default_payload["metadata"]["year"] = year
        default_payload["metadata"]["month"] = month
        return apply_shared_correction_settings(hydrate_payload_corrections(default_payload))

    with data_file.open("r", encoding="utf-8") as file:
        data = json.load(file)

    return apply_shared_correction_settings(
        hydrate_payload_corrections(
            merge_payload_with_saved_data(data, equipment_code=equipment_code, form_key=form_key)
        )
    )


def get_comparable_payload(payload: dict[str, Any]) -> dict[str, Any]:
    comparable_payload = json.loads(json.dumps(payload, ensure_ascii=False))
    comparable_payload.pop("change_log", None)
    return comparable_payload


def get_payload_signature(payload: dict[str, Any]) -> str:
    return json.dumps(get_comparable_payload(payload), ensure_ascii=False, sort_keys=True)


def remember_saved_snapshot(payload: dict[str, Any]) -> None:
    st.session_state["last_saved_payload_snapshot"] = get_comparable_payload(payload)
    st.session_state["last_saved_payload_signature"] = get_payload_signature(payload)


def format_change_log_item(payload: dict[str, Any], metric: dict[str, Any], value: str) -> str:
    formatted = format_metric_value(payload, metric, value)
    return formatted or "vacio"


def build_change_log_entry(previous_payload: dict[str, Any], current_payload: dict[str, Any]) -> dict[str, Any] | None:
    definition = get_form_definition(current_payload["metadata"]["form_key"])
    items: list[str] = []

    for day in range(1, 32):
        day_key = str(day)
        previous_record = previous_payload.get("daily_records", {}).get(day_key, {})
        current_record = current_payload.get("daily_records", {}).get(day_key, {})

        for metric in definition["metrics"]:
            metric_label = get_primary_metric_display_label(current_payload, metric)
            previous_values = previous_record.get(metric["key"], ["", "", ""])
            current_values = current_record.get(metric["key"], ["", "", ""])
            for index, slot_label in enumerate(TIME_SLOTS):
                previous_value = str(previous_values[index] if index < len(previous_values) else "").strip()
                current_value = str(current_values[index] if index < len(current_values) else "").strip()
                if not previous_value or previous_value == current_value:
                    continue
                old_display = format_change_log_item(current_payload, metric, previous_value)
                new_display = format_change_log_item(current_payload, metric, current_value)
                items.append(f"Cambio: Dia {day} {metric_label} {slot_label}: {old_display} -> {new_display}")

        previous_verified = str(previous_record.get("verified_by", "")).strip()
        current_verified = str(current_record.get("verified_by", "")).strip()
        if previous_verified and previous_verified != current_verified:
            items.append(f"Cambio: Dia {day} Verifico: {previous_verified} -> {current_verified or 'vacio'}")

    if not items:
        return None

    return {
        "timestamp": get_local_now().isoformat(timespec="seconds"),
        "user": str(st.session_state.get("usuario_email", "")).strip(),
        "items": items,
    }


def compose_observations_export_text(payload: dict[str, Any]) -> str:
    base_observations = str(payload["monthly_closure"].get("observations", "")).strip()
    non_working_days = sorted(int(day) for day in payload.get("non_working_days", []))
    change_log = payload.get("change_log", [])
    sections: list[str] = []

    if base_observations:
        sections.append(base_observations)

    if non_working_days:
        days_text = ", ".join(str(day) for day in non_working_days)
        sections.append(f"* Los dias {days_text} no aplican por marcarse como no laborados.")

    canceled_lines: list[str] = []
    for day in range(1, 32):
        record = payload.get("daily_records", {}).get(str(day), {})
        if not record.get("active", False):
            continue
        canceled_slots = normalize_canceled_slots(record.get("canceled_slots", []))
        if not canceled_slots:
            continue
        if len(canceled_slots) >= len(TIME_SLOTS):
            slots_text = "dia completo"
        else:
            slots_text = ", ".join(TIME_SLOTS[index] for index in canceled_slots)
        note = str(record.get("cancellation_note", "")).strip()
        suffix = f": {note}" if note else ""
        canceled_lines.append(f"* Dia {day}: captura cancelada en {slots_text}{suffix}.")
    if canceled_lines:
        sections.extend(canceled_lines)

    if not change_log:
        return "\n\n".join(sections).strip()

    audit_lines: list[str] = []
    for entry in change_log[-6:]:
        items = entry.get("items", [])
        if not items:
            continue
        visible_items = items[:3]
        detail = "; ".join(str(item) for item in visible_items)
        if len(items) > len(visible_items):
            detail += f"; y {len(items) - len(visible_items)} cambio(s) mas"
        audit_lines.append(detail)

    if not audit_lines:
        return "\n\n".join(sections).strip()

    audit_block = "Registro de correcciones:\n" + "\n".join(audit_lines)
    sections.append(audit_block)
    return "\n\n".join(section for section in sections if section).strip()


def save_payload(payload: dict[str, Any]) -> str:
    previous_snapshot = st.session_state.get("last_saved_payload_snapshot")
    if isinstance(previous_snapshot, dict):
        change_entry = build_change_log_entry(previous_snapshot, get_comparable_payload(payload))
        if change_entry is not None:
            payload.setdefault("change_log", []).append(change_entry)
            payload["change_log"] = payload["change_log"][-80:]

    try:
        save_shared_correction_settings(payload)
    except Exception:
        pass

    if supabase_storage_enabled():
        try:
            save_remote_period_payload(
                payload,
                updated_by=str(st.session_state.get("usuario_email", "")).strip(),
            )
            remember_saved_snapshot(payload)
            return "Supabase"
        except Exception:
            pass

    data_file = get_period_file(payload)
    with data_file.open("w", encoding="utf-8") as file:
        json.dump(payload, file, ensure_ascii=False, indent=2)
    remember_saved_snapshot(payload)
    return "Local JSON"


def get_supabase_status() -> tuple[bool, str]:
    if supabase_storage_enabled():
        return True, "Supabase"
    return False, "Local JSON"


def get_row_group(payload: dict[str, Any], day: int) -> dict[str, int]:
    layout = get_form_definition(payload["metadata"]["form_key"])["layout"]
    return layout["top"] if day <= 16 else layout["bottom"]


def get_day_column(day: int, slot_index: int) -> int:
    if day <= 16:
        return DAY_BLOCK_START_COLUMNS[day] + slot_index
    return DAY_BLOCK_START_COLUMNS[day] + slot_index


def normalize_excel_date(date_value: str) -> str:
    if not date_value:
        return ""

    try:
        parsed = datetime.fromisoformat(date_value)
        return parsed.strftime("%d/%m/%Y")
    except ValueError:
        return date_value


def get_effective_record_date(record: dict[str, Any]) -> str:
    recorded_on = record.get("recorded_on", "").strip()
    if recorded_on:
        return normalize_excel_date(recorded_on)
    return ""


def get_format_specific_copy(payload: dict[str, Any]) -> dict[str, str]:
    form_key = payload["metadata"]["form_key"]
    if form_key == "incubadoras":
        return {
            "config_intro": "Este formato registra temperatura y %CO2. La corrección aplica solo a la temperatura.",
            "daily_intro": "Cada día laborado captura temperatura, %CO2, tres responsables por horario y un verificador del bloque.",
            "notes_label": "Incidencia del día (opcional)",
        }
    if form_key == "condiciones_ambientales":
        return {
            "config_intro": "Este formato monitorea condiciones ambientales con un thermohigrómetro y rangos de corrección específicos por hoja.",
            "daily_intro": "Cada día laborado registra lecturas ambientales por horario, responsables, verificación y fecha del bloque.",
            "notes_label": "Incidencia ambiental del día (opcional)",
        }
    if form_key == "ultracongeladores":
        return {
            "config_intro": "Este formato controla ultracongeladores. Algunos equipos usan factores reales y otros están marcados como N/A.",
            "daily_intro": "Cada día laborado captura temperatura por horario, responsables, verificación y fecha del bloque.",
            "notes_label": "Incidencia del ultracongelador (opcional)",
        }
    if form_key == "refrigeradores":
        return {
            "config_intro": "Este formato controla refrigeradores con rangos y factores específicos por equipo.",
            "daily_intro": "Cada día laborado captura temperatura por horario, responsables, verificación y fecha del bloque.",
            "notes_label": "Incidencia del refrigerador (opcional)",
        }
    return {
        "config_intro": "Este formato controla congeladores con rangos y factores específicos por equipo.",
        "daily_intro": "Cada día laborado captura temperatura por horario, responsables, verificación y fecha del bloque.",
        "notes_label": "Incidencia del congelador (opcional)",
    }


def format_percentage_display(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    if "%" in text:
        return text

    normalized = text.replace(",", ".")
    try:
        numeric = float(normalized)
    except ValueError:
        return text

    percentage_value = numeric * 100 if abs(numeric) <= 1 else numeric
    if percentage_value.is_integer():
        return f"{int(percentage_value)}%"
    return f"{percentage_value:.2f}".rstrip("0").rstrip(".") + "%"


def is_ambient_form(payload: dict[str, Any]) -> bool:
    return payload["metadata"]["form_key"] == "condiciones_ambientales"


def is_ambient_humidity_payload(payload: dict[str, Any]) -> bool:
    equipment_code = str(payload["metadata"].get("equipment_code", "")).upper()
    return is_ambient_form(payload) and equipment_code.startswith("HUMEDAD")


def get_ambient_variable_name(payload: dict[str, Any]) -> str:
    return "% Humedad" if is_ambient_humidity_payload(payload) else "Temperatura"


def get_primary_metric_display_label(payload: dict[str, Any], metric: dict[str, Any]) -> str:
    if is_ambient_form(payload):
        return "Humedad medida" if is_ambient_humidity_payload(payload) else "Temperatura medida"
    return str(metric["label"])


def is_humidity_metric(payload: dict[str, Any], metric: dict[str, Any]) -> bool:
    return is_ambient_humidity_payload(payload) and metric["key"] == "measured_temperatures"


def is_incubator_co2_metric(payload: dict[str, Any], metric: dict[str, Any]) -> bool:
    return (
        payload["metadata"]["form_key"] == "incubadoras"
        and metric["key"] == "secondary_measurements"
    )


def parse_measurement_number(raw_value: str) -> float | None:
    cleaned = (
        raw_value.strip()
        .replace(",", ".")
        .replace("%", "")
        .replace("≤", "")
        .replace("≥", "")
    )
    if not cleaned:
        return None

    try:
        numeric = float(cleaned)
    except ValueError:
        return None

    return numeric * 100 if 0 < abs(numeric) <= 1 else numeric


def format_decimal_value(value: float, decimal_places: int = TEMPERATURE_DECIMAL_PLACES) -> str:
    formatted = f"{value:.{decimal_places}f}"
    if "." not in formatted:
        return formatted
    return formatted.rstrip("0").rstrip(".")


def format_metric_value(payload: dict[str, Any], metric: dict[str, Any], value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    if is_humidity_metric(payload, metric):
        return format_percentage_display(text)
    if is_incubator_co2_metric(payload, metric):
        return format_percentage_display(text)
    numeric_value = parse_measurement_number(text)
    if numeric_value is not None:
        return format_decimal_value(numeric_value)
    return text


def render_sidebar(
    payload: dict[str, Any],
    form_keys: list[str],
    equipment_codes: list[str],
) -> tuple[str, str]:
    st.sidebar.title("Formatos")
    st.sidebar.write(f"Sesion: `{st.session_state.get('usuario_email', '')}`")
    st.sidebar.write(f"Perfil: `{st.session_state.get('rol_usuario', 'captura')}`")
    display_name = get_current_user_display_name()
    if display_name:
        st.sidebar.write(f"Firma automatica: `{display_name}`")
    if st.sidebar.button("Cerrar sesion", use_container_width=True):
        st.session_state["autenticado"] = False
        st.session_state["usuario_email"] = ""
        st.session_state["usuario_nombre"] = ""
        st.session_state["es_admin"] = False
        st.session_state["rol_usuario"] = "captura"
        st.rerun()

    selected_form_key = st.sidebar.selectbox(
        "Formato",
        options=form_keys,
        index=form_keys.index(payload["metadata"]["form_key"]),
        format_func=lambda key: FORM_DEFINITIONS[key]["label"],
        key="sidebar_form_key",
    )
    selected_equipment = st.sidebar.selectbox(
        "Equipo",
        options=equipment_codes,
        index=equipment_codes.index(payload["metadata"]["equipment_code"]),
        key="sidebar_equipment",
    )
    st.sidebar.caption(payload["metadata"]["form_label"])
    st.sidebar.write(
        f"Mes de trabajo: `{MONTHS[payload['metadata']['month']]} {payload['metadata']['year']}`"
    )
    if not is_capture_role():
        st.sidebar.write(f"Laboratorio: `{payload['metadata']['laboratory']}`")
        st.sidebar.write(f"Equipo / instrumento: `{payload['metadata']['equipment_name']}`")
    render_user_admin_sidebar()
    return selected_form_key, selected_equipment


def render_configuration(payload: dict[str, Any]) -> None:
    if not can_edit_sensitive_configuration():
        st.subheader("Captura rapida")
        st.info(
            "Modo captura: registra el dia sugerido, el bloque correspondiente y guarda. "
            "Tu nombre se completa automaticamente en Realizo cuando captures una lectura."
        )
        return

    st.subheader("1. Configuracion del mes")
    metadata = payload["metadata"]
    correction_bands = payload["correction_bands"]
    correction_factors = payload["correction_factors"]
    correction_operations = payload["correction_operations"]
    period_key = st.session_state.get("period_key", get_period_key(payload))
    copy = get_format_specific_copy(payload)
    allow_sensitive_edits = can_edit_sensitive_configuration()
    allow_correction_edits = can_edit_correction_settings()

    month_col, year_col = st.columns(2)
    metadata["month"] = month_col.selectbox(
        "Mes",
        options=list(MONTHS.keys()),
        index=list(MONTHS.keys()).index(metadata["month"]),
        format_func=lambda value: MONTHS[value],
    )
    metadata["year"] = year_col.number_input(
        "Año",
        min_value=2024,
        max_value=2100,
        value=int(metadata["year"]),
        step=1,
    )
    st.caption(copy["config_intro"])
    if allow_sensitive_edits:
        st.caption("Edicion sensible habilitada para este perfil.")
    else:
        st.caption("Rangos, inventario y metadata del equipo estan en solo lectura para este perfil.")

    definition = get_form_definition(metadata["form_key"])
    is_incubator_form = metadata["form_key"] == "incubadoras"
    ambient_variable_name = get_ambient_variable_name(payload) if is_ambient_form(payload) else ""
    is_ambient_humidity = is_ambient_humidity_payload(payload)
    summary_cols = st.columns(4)
    equipment_name_key = f"equipment_name_{period_key}"
    if equipment_name_key not in st.session_state:
        st.session_state[equipment_name_key] = str(metadata["equipment_name"])
    metadata["equipment_name"] = summary_cols[0].text_input(
        "Equipo / instrumento",
        key=equipment_name_key,
        disabled=not allow_sensitive_edits,
    )
    brand_key = f"brand_{period_key}"
    if brand_key not in st.session_state:
        st.session_state[brand_key] = str(metadata["brand"])
    metadata["brand"] = summary_cols[1].text_input(
        "Marca",
        key=brand_key,
        disabled=not allow_sensitive_edits,
    )
    model_key = f"model_{period_key}"
    if model_key not in st.session_state:
        st.session_state[model_key] = str(metadata["model"])
    metadata["model"] = summary_cols[2].text_input(
        "Modelo",
        key=model_key,
        disabled=not allow_sensitive_edits,
    )
    inventory_key = f"inventory_{period_key}"
    if inventory_key not in st.session_state:
        st.session_state[inventory_key] = str(metadata["inventory_code"])
    metadata["inventory_code"] = summary_cols[3].text_input(
        "Inventario / código",
        key=inventory_key,
        disabled=not allow_sensitive_edits,
    )

    detail_cols = st.columns(4)
    serial_key = f"serial_{period_key}"
    if serial_key not in st.session_state:
        st.session_state[serial_key] = str(metadata["serial_number"])
    metadata["serial_number"] = detail_cols[0].text_input(
        "Serie",
        key=serial_key,
        disabled=not allow_sensitive_edits,
    )
    normal_key = f"normal_{period_key}"
    if normal_key not in st.session_state:
        st.session_state[normal_key] = (
            format_percentage_display(metadata["temperature_label"])
            if is_ambient_humidity
            else str(metadata["temperature_label"])
        )
    metadata["temperature_label"] = detail_cols[1].text_input(
        f"{ambient_variable_name} normal" if ambient_variable_name else "Valor normal",
        key=normal_key,
        disabled=not allow_sensitive_edits,
    )
    minimum_key = f"minimum_{period_key}"
    if minimum_key not in st.session_state:
        st.session_state[minimum_key] = (
            format_percentage_display(metadata["minimum_label"])
            if is_ambient_humidity
            else str(metadata["minimum_label"])
        )
    metadata["minimum_label"] = detail_cols[2].text_input(
        f"Mínima {ambient_variable_name.lower()}" if ambient_variable_name else "Mínima",
        key=minimum_key,
        disabled=not allow_sensitive_edits,
    )
    maximum_key = f"maximum_{period_key}"
    if maximum_key not in st.session_state:
        st.session_state[maximum_key] = (
            format_percentage_display(metadata["maximum_label"])
            if is_ambient_humidity
            else str(metadata["maximum_label"])
        )
    metadata["maximum_label"] = detail_cols[3].text_input(
        f"Máxima {ambient_variable_name.lower()}" if ambient_variable_name else "Máxima",
        key=maximum_key,
        disabled=not allow_sensitive_edits,
    )

    if metadata.get("secondary_label"):
        secondary_section_title = "%CO2 visible" if is_incubator_form else "Variable secundaria visible"
        secondary_primary_label = "%CO2 normal" if is_incubator_form else "Nombre de la segunda variable"
        secondary_minimum_label = "Mínima %CO2" if is_incubator_form else "Mínima secundaria"
        secondary_maximum_label = "Máxima %CO2" if is_incubator_form else "Máxima secundaria"
        st.markdown(f"**{secondary_section_title}**")
        secondary_cols = st.columns(3)
        secondary_label_key = f"secondary_label_{period_key}"
        if secondary_label_key not in st.session_state:
            st.session_state[secondary_label_key] = (
                format_percentage_display(metadata["secondary_label"])
                if is_incubator_form
                else str(metadata["secondary_label"])
            )
        elif is_incubator_form:
            st.session_state[secondary_label_key] = format_percentage_display(st.session_state[secondary_label_key])
        metadata["secondary_label"] = secondary_cols[0].text_input(
            secondary_primary_label,
            key=secondary_label_key,
            disabled=not allow_sensitive_edits,
        )
        secondary_minimum_key = f"secondary_minimum_{period_key}"
        if secondary_minimum_key not in st.session_state:
            st.session_state[secondary_minimum_key] = (
                format_percentage_display(metadata.get("secondary_minimum_label", ""))
                if is_incubator_form
                else str(metadata.get("secondary_minimum_label", ""))
            )
        elif is_incubator_form:
            st.session_state[secondary_minimum_key] = format_percentage_display(st.session_state[secondary_minimum_key])
        metadata["secondary_minimum_label"] = secondary_cols[1].text_input(
            secondary_minimum_label,
            key=secondary_minimum_key,
            disabled=not allow_sensitive_edits,
        )
        secondary_maximum_key = f"secondary_maximum_{period_key}"
        if secondary_maximum_key not in st.session_state:
            st.session_state[secondary_maximum_key] = (
                format_percentage_display(metadata.get("secondary_maximum_label", ""))
                if is_incubator_form
                else str(metadata.get("secondary_maximum_label", ""))
            )
        elif is_incubator_form:
            st.session_state[secondary_maximum_key] = format_percentage_display(st.session_state[secondary_maximum_key])
        metadata["secondary_maximum_label"] = secondary_cols[2].text_input(
            secondary_maximum_label,
            key=secondary_maximum_key,
            disabled=not allow_sensitive_edits,
        )

    if definition["supports_corrections"] and correction_factors:
        if allow_correction_edits:
            st.caption("Rangos, operacion y factor de correccion editables para admins.")
        else:
            st.caption("Los rangos y factores de correccion se muestran en solo lectura. Solo admins pueden modificarlos.")
        factor_labels = list(correction_factors.keys())
        ranges_per_row = 3 if is_ambient_form(payload) else min(3, max(1, len(factor_labels)))
        for start_index in range(0, len(factor_labels), ranges_per_row):
            row_keys = factor_labels[start_index:start_index + ranges_per_row]
            factor_cols = st.columns(len(row_keys))
            for factor_col, factor_key in zip(factor_cols, row_keys):
                band = dict(correction_bands.get(factor_key, {}))
                label = str(band.get("label", factor_key.replace("_", " ").title()))
                title_prefix = ambient_variable_name if is_ambient_form(payload) else ("Temperatura" if is_incubator_form else "Rango")
                factor_col.markdown(f"**{title_prefix} {label}**")

                min_key = f"range_min_{period_key}_{factor_key}"
                max_key = f"range_max_{period_key}_{factor_key}"
                if min_key not in st.session_state:
                    st.session_state[min_key] = float(band.get("min", 0.0))
                if max_key not in st.session_state:
                    st.session_state[max_key] = float(band.get("max", 0.0))

                min_col, max_col = factor_col.columns(2)
                min_value = min_col.number_input(
                    "Minimo",
                    value=float(st.session_state[min_key]),
                    step=0.001,
                    format="%.3f",
                    key=min_key,
                    disabled=not allow_correction_edits,
                )
                max_value = max_col.number_input(
                    "Maximo",
                    value=float(st.session_state[max_key]),
                    step=0.001,
                    format="%.3f",
                    key=max_key,
                    disabled=not allow_correction_edits,
                )
                separator = str(band.get("separator", infer_range_separator(label))).strip() or "-"
                correction_bands[factor_key] = {
                    "label": build_range_label(min_value, max_value, separator),
                    "min": float(min_value),
                    "max": float(max_value),
                    "separator": separator,
                }

                operation_col, value_col = factor_col.columns([1, 2])
                correction_operations[factor_key] = operation_col.selectbox(
                    "Operacion",
                    options=["+", "-"],
                    index=0 if correction_operations[factor_key] == "+" else 1,
                    key=f"operation_{period_key}_{factor_key}",
                    disabled=not allow_correction_edits,
                )
                correction_factors[factor_key] = value_col.number_input(
                    "Factor de correccion",
                    value=float(correction_factors[factor_key]),
                    step=0.001,
                    format="%.3f",
                    key=f"factor_{period_key}_{factor_key}",
                    disabled=not allow_correction_edits,
                )
    else:
        st.caption(
            "Este equipo no usa factores de corrección editables en la plantilla o están marcados como N/A."
        )


def render_non_working_days(payload: dict[str, Any]) -> None:
    st.subheader("2. Dias no laborados")
    st.caption(
        "Marca los dias que no aplican para la toma. Esta seleccion se comparte con todos los formatos "
        "del mismo mes y ano."
    )
    allow_schedule_edits = can_edit_schedule()
    if not allow_schedule_edits:
        st.caption("Este apartado esta en solo lectura para tu perfil.")
    year = int(payload["metadata"]["year"])
    month = int(payload["metadata"]["month"])
    shared_days = load_shared_non_working_days(year, month)
    if shared_days is not None:
        payload["non_working_days"] = shared_days
    days = list(range(1, 32))
    selector_key = f"global_non_working_days_{year}_{month:02d}"
    if selector_key not in st.session_state:
        st.session_state[selector_key] = list(sorted(int(day) for day in payload["non_working_days"]))

    if hasattr(st, "pills"):
        selected_days = st.pills(
            "Selecciona dias no laborados",
            options=days,
            default=st.session_state[selector_key],
            selection_mode="multi",
            key=selector_key,
            disabled=not allow_schedule_edits,
        )
    else:
        selected_days = st.multiselect(
            "Selecciona dias no laborados",
            options=days,
            default=st.session_state[selector_key],
            key=selector_key,
            disabled=not allow_schedule_edits,
        )

    payload["non_working_days"] = sorted(int(day) for day in (selected_days or []))
    if allow_schedule_edits:
        save_shared_non_working_days(year, month, payload["non_working_days"])

    if payload["non_working_days"]:
        st.caption(
            "Dias marcados como no laborados: "
            + ", ".join(str(day) for day in payload["non_working_days"])
        )
    else:
        st.caption("No hay dias marcados como no laborados.")

    for day in range(1, 32):
        payload["daily_records"][str(day)]["active"] = day not in payload["non_working_days"]


def render_daily_capture(payload: dict[str, Any]) -> None:
    st.subheader("3. Captura diaria")
    copy = get_format_specific_copy(payload)
    st.caption(copy["daily_intro"])
    allow_daily_edits = can_edit_daily_records()
    allow_verification_edits = can_verify_daily_records()
    if not allow_daily_edits:
        st.caption("La captura diaria esta en solo lectura para tu perfil.")
    period_key = st.session_state.get("period_key", get_period_key(payload))
    definition = get_form_definition(payload["metadata"]["form_key"])
    metrics = definition["metrics"]

    active_days = [
        day for day in range(1, 32) if payload["daily_records"][str(day)]["active"]
    ]
    if not active_days:
        st.info("No hay dias activos. Marca al menos un dia laborado para capturar informacion.")
        return

    preferred_day = get_preferred_capture_day(payload, active_days)
    current_slot_index = get_current_time_slot_index() if is_current_period(payload) else None
    ordered_active_days = get_ordered_active_days(payload, active_days, preferred_day)
    if is_current_period(payload):
        slot_text = TIME_SLOTS[current_slot_index] if current_slot_index is not None else "fuera de horario de captura"
        st.info(f"Dia sugerido por el sistema: {preferred_day}. Bloque actual: {slot_text}.")
    days_to_render = ordered_active_days
    if is_capture_role():
        selected_capture_day = st.selectbox(
            "Dia a capturar",
            options=ordered_active_days,
            index=0,
            format_func=lambda value: f"Dia {value}",
            key=f"capture_day_picker_{period_key}",
        )
        days_to_render = [int(selected_capture_day)]
        st.caption("Para captura se muestra un solo dia a la vez. Si necesitas corregir otro dia, seleccionalo aqui.")

    for day in days_to_render:
        record = payload["daily_records"][str(day)]
        with st.expander(f"Dia {day}", expanded=is_capture_role() or day == preferred_day):
            record["canceled_slots"] = normalize_canceled_slots(record.get("canceled_slots", []))
            cancellation_cols = st.columns([1, 2, 3])
            full_cancel_key = f"cancel_full_day_{period_key}_{day}"
            full_day_canceled = cancellation_cols[0].checkbox(
                "Cancelar dia completo",
                value=is_day_fully_canceled(record),
                key=full_cancel_key,
                disabled=not allow_daily_edits,
            )
            if full_day_canceled:
                record["canceled_slots"] = list(range(len(TIME_SLOTS)))
                cancellation_cols[1].caption("Las tres capturas del dia se exportaran como N/A.")
            else:
                canceled_key = f"canceled_slots_{period_key}_{day}"
                canceled_default = [] if is_day_fully_canceled(record) and not st.session_state.get(full_cancel_key, False) else record["canceled_slots"]
                selected_canceled_slots = cancellation_cols[1].multiselect(
                    "Horarios cancelados",
                    options=list(range(len(TIME_SLOTS))),
                    default=canceled_default,
                    format_func=lambda index: TIME_SLOTS[index],
                    key=canceled_key,
                    disabled=not allow_daily_edits,
                )
                record["canceled_slots"] = normalize_canceled_slots(selected_canceled_slots)

            note_key = f"cancellation_note_{period_key}_{day}"
            if note_key not in st.session_state:
                st.session_state[note_key] = str(record.get("cancellation_note", ""))
            record["cancellation_note"] = cancellation_cols[2].text_input(
                "Motivo de cancelacion",
                key=note_key,
                disabled=not allow_daily_edits or not record["canceled_slots"],
                placeholder="Ej. salida temprana, mantenimiento, falla electrica",
            )

            captured_slot_flags = [False, False, False]
            for metric in metrics:
                metric_label = get_primary_metric_display_label(payload, metric)
                if len(metrics) > 1 or is_ambient_form(payload):
                    if metric["key"] == "measured_temperatures":
                        if is_ambient_form(payload):
                            heading = get_ambient_variable_name(payload)
                        elif payload["metadata"]["form_key"] == "incubadoras":
                            heading = "Temperatura medida"
                        else:
                            heading = payload["metadata"].get("temperature_label", "Temperatura")
                    else:
                        heading = "%CO2" if payload["metadata"]["form_key"] == "incubadoras" else payload["metadata"].get("secondary_label", "Variable secundaria")
                    st.markdown(f"**{heading}**")
                metric_cols = st.columns(3)
                metric_values = list(record.get(metric["key"], ["", "", ""]))
                while len(metric_values) < len(TIME_SLOTS):
                    metric_values.append("")
                corrected_values = list(record.get("corrected_temperatures", ["", "", ""]))
                while len(corrected_values) < len(TIME_SLOTS):
                    corrected_values.append("")
                for index, label in enumerate(TIME_SLOTS):
                    input_key = f"{metric['key']}_{period_key}_{day}_{index}"
                    slot_label = f"{metric_label} {label}"
                    if day == preferred_day and index == current_slot_index:
                        slot_label = f"{slot_label} (bloque actual)"
                    if is_slot_canceled(record, index):
                        metric_cols[index].text_input(
                            slot_label,
                            value="N/A",
                            key=f"{input_key}_canceled_display",
                            disabled=True,
                        )
                        if metric.get("corrected", False):
                            corrected_values[index] = "N/A"
                            metric_cols[index].caption("Cancelada: N/A")
                        continue
                    if input_key not in st.session_state:
                        st.session_state[input_key] = metric_values[index]
                    metric_values[index] = metric_cols[index].text_input(
                        slot_label,
                        key=input_key,
                        disabled=not allow_daily_edits,
                        placeholder="-20.123" if metric["unit"] == "°C" else "",
                    )
                    if metric_values[index].strip():
                        captured_slot_flags[index] = True
                    if metric.get("corrected", False):
                        corrected_value = calculate_corrected_temperature(
                            metric_values[index],
                            payload["correction_bands"],
                            payload["correction_factors"],
                            payload["correction_operations"],
                        )
                        corrected_values[index] = corrected_value
                        corrected_display = format_metric_value(payload, metric, corrected_value)
                        metric_cols[index].caption(
                            f"Corregida: {corrected_display}" if corrected_display else "Corregida: pendiente"
                        )
                record[metric["key"]] = metric_values
                if metric["key"] == "measured_temperatures":
                    record["corrected_temperatures"] = corrected_values

            actor_cols = st.columns(3)
            performed_by_slots = list(record.get("performed_by_slots", ["", "", ""]))
            while len(performed_by_slots) < len(TIME_SLOTS):
                performed_by_slots.append("")
            automatic_signature_name = get_current_user_display_name()
            for index, label in enumerate(TIME_SLOTS):
                input_key = f"performed_{period_key}_{day}_{index}"
                performed_label = f"Realizo {label}"
                if day == preferred_day and index == current_slot_index:
                    performed_label = f"{performed_label} (bloque actual)"
                if is_slot_canceled(record, index):
                    actor_cols[index].text_input(
                        performed_label,
                        value="N/A",
                        key=f"{input_key}_canceled_display",
                        disabled=True,
                    )
                    actor_cols[index].caption("Captura cancelada")
                    continue
                if captured_slot_flags[index] and automatic_signature_name and not performed_by_slots[index].strip():
                    performed_by_slots[index] = automatic_signature_name
                    st.session_state[input_key] = automatic_signature_name
                if input_key not in st.session_state:
                    st.session_state[input_key] = performed_by_slots[index]
                performed_by_slots[index] = actor_cols[index].text_input(
                    performed_label,
                    key=input_key,
                    disabled=not allow_daily_edits,
                )
                performed_signature_name = get_signature_display_name(performed_by_slots[index])
                if performed_by_slots[index].strip():
                    if performed_signature_name:
                        actor_cols[index].caption(f"Firma detectada: {performed_signature_name}")
                    else:
                        actor_cols[index].caption("Firma digital: sin coincidencia")
                else:
                    actor_cols[index].caption("Firma digital: pendiente")
            record["performed_by_slots"] = performed_by_slots

            verifier_cols = st.columns(3)
            verified_key = f"verified_{period_key}_{day}"
            if verified_key not in st.session_state:
                st.session_state[verified_key] = record["verified_by"]
            record["verified_by"] = verifier_cols[0].text_input(
                "Verifico bloque del dia",
                key=verified_key,
                disabled=not allow_verification_edits or is_day_fully_canceled(record),
            )
            verified_signature_name = get_signature_display_name(record["verified_by"])
            if record["verified_by"].strip():
                if verified_signature_name:
                    verifier_cols[0].caption(f"Firma detectada: {verified_signature_name}")
                else:
                    verifier_cols[0].caption("Firma digital: sin coincidencia")
            else:
                verifier_cols[0].caption("Firma digital: pendiente")
            current_date = get_period_default_date(payload, day, record)
            date_key = f"date_{period_key}_{day}"
            if date_key not in st.session_state:
                st.session_state[date_key] = current_date
            elif should_follow_live_record_date(payload, record):
                st.session_state[date_key] = current_date
            original_mode = str(record.get("recorded_on_mode", "auto")).strip().lower() or "auto"
            selected_date = verifier_cols[1].date_input(
                "Fecha de verificacion",
                min_value=date(
                    int(payload["metadata"]["year"]),
                    int(payload["metadata"]["month"]),
                    1,
                ),
                max_value=date(
                    int(payload["metadata"]["year"]),
                    int(payload["metadata"]["month"]),
                    monthrange(
                        int(payload["metadata"]["year"]),
                        int(payload["metadata"]["month"]),
                    )[1],
                ),
                key=date_key,
                disabled=not allow_verification_edits or is_day_fully_canceled(record),
            )
            record["recorded_on"] = selected_date.isoformat()
            if str(record["verified_by"]).strip():
                record["recorded_on_mode"] = "manual"
            elif selected_date != current_date:
                record["recorded_on_mode"] = "manual"
            elif should_follow_live_record_date(payload, record):
                record["recorded_on_mode"] = "auto"
            else:
                record["recorded_on_mode"] = original_mode

            notes_key = f"notes_{period_key}_{day}"
            if notes_key not in st.session_state:
                st.session_state[notes_key] = record.get("notes", "")
            record["notes"] = st.text_input(
                copy["notes_label"],
                key=notes_key,
                disabled=not allow_daily_edits,
            )


def parse_streamlit_date(value: str) -> date:
    try:
        return datetime.fromisoformat(value).date()
    except ValueError:
        return get_local_now().date()


def is_current_period(payload: dict[str, Any]) -> bool:
    today = get_local_now().date()
    return (
        int(payload["metadata"]["year"]) == today.year
        and int(payload["metadata"]["month"]) == today.month
    )


def get_preferred_capture_day(payload: dict[str, Any], active_days: list[int]) -> int:
    today = get_local_now().date()
    if is_current_period(payload) and today.day in active_days:
        return today.day
    return active_days[0]


def get_current_time_slot_index() -> int | None:
    now_time = get_local_now().time()
    slot_ranges = [
        (7, 10),
        (11, 14),
        (15, 18),
    ]
    for index, (start_hour, end_hour) in enumerate(slot_ranges):
        if start_hour <= now_time.hour < end_hour:
            return index
    return None


def get_ordered_active_days(payload: dict[str, Any], active_days: list[int], preferred_day: int) -> list[int]:
    if not active_days:
        return []
    if is_current_period(payload) and preferred_day in active_days:
        return [preferred_day, *[day for day in active_days if day != preferred_day]]
    return active_days


def get_row_period_date(payload: dict[str, Any], day: int) -> date:
    year = int(payload["metadata"]["year"])
    month = int(payload["metadata"]["month"])
    last_day = monthrange(year, month)[1]
    return date(year, month, min(day, last_day))


def should_follow_live_record_date(payload: dict[str, Any], record: dict[str, Any]) -> bool:
    mode = str(record.get("recorded_on_mode", "auto")).strip().lower() or "auto"
    if mode != "auto":
        return False
    if not is_current_period(payload):
        return False
    if str(record.get("verified_by", "")).strip():
        return False
    return True


def get_period_default_date(payload: dict[str, Any], day: int, record: dict[str, Any]) -> date:
    year = int(payload["metadata"]["year"])
    month = int(payload["metadata"]["month"])
    recorded_on = str(record.get("recorded_on", "")).strip()
    if should_follow_live_record_date(payload, record):
        return date.today()

    if recorded_on.strip():
        parsed = parse_streamlit_date(recorded_on)
        if parsed.year == year and parsed.month == month:
            return parsed

    return get_row_period_date(payload, day)


def get_period_end_date(payload: dict[str, Any]) -> date:
    year = int(payload["metadata"]["year"])
    month = int(payload["metadata"]["month"])
    last_day = monthrange(year, month)[1]
    return date(year, month, last_day)


def uses_default_closure_date(payload: dict[str, Any], reviewed_on: str) -> bool:
    if not reviewed_on.strip():
        return True
    parsed = parse_streamlit_date(reviewed_on)
    return parsed == get_period_end_date(payload)


def get_closure_default_date(payload: dict[str, Any]) -> date:
    closure = payload["monthly_closure"]
    reviewed_on = str(closure.get("reviewed_on", "")).strip()
    reviewed_by = str(closure.get("reviewed_by", "")).strip()

    if is_current_period(payload) and uses_default_closure_date(payload, reviewed_on):
        return date.today()

    if reviewed_on:
        parsed = parse_streamlit_date(reviewed_on)
        if parsed.year == int(payload["metadata"]["year"]) and parsed.month == int(payload["metadata"]["month"]):
            return parsed

    if reviewed_by:
        return parse_streamlit_date(reviewed_on)

    return get_period_end_date(payload)


def render_monthly_closure(payload: dict[str, Any]) -> None:
    st.subheader("4. Cierre del formato")
    closure = payload["monthly_closure"]
    period_key = st.session_state.get("period_key", get_period_key(payload))
    allow_closure_edits = can_close_period()
    if not allow_closure_edits:
        st.caption("El cierre del formato esta en solo lectura para tu perfil.")

    observations_key = f"observations_{period_key}"
    if observations_key not in st.session_state:
        st.session_state[observations_key] = closure["observations"]
    closure["observations"] = st.text_area(
        "Observaciones",
        key=observations_key,
        height=120,
        placeholder="Anota incidencias, mantenimiento, ajustes o aclaraciones del periodo.",
        disabled=not allow_closure_edits,
    )
    if payload.get("change_log"):
        st.caption("Las correcciones sobre capturas previas se anexan automaticamente en Observaciones al exportar.")
    review_cols = st.columns(2)
    reviewed_by_key = f"reviewed_by_{period_key}"
    if reviewed_by_key not in st.session_state:
        st.session_state[reviewed_by_key] = closure["reviewed_by"]
    closure["reviewed_by"] = review_cols[0].text_input(
        "Reviso",
        key=reviewed_by_key,
        disabled=not allow_closure_edits,
    )
    reviewed_signature_name = get_signature_display_name(closure["reviewed_by"])
    if closure["reviewed_by"].strip():
        if reviewed_signature_name:
            review_cols[0].caption(f"Firma detectada: {reviewed_signature_name}")
        else:
            review_cols[0].caption("Firma digital: sin coincidencia")
    else:
        review_cols[0].caption("Firma digital: pendiente")
    reviewed_on_key = f"reviewed_on_{period_key}"
    default_review_date = get_closure_default_date(payload)
    if reviewed_on_key not in st.session_state:
        st.session_state[reviewed_on_key] = default_review_date
    elif is_current_period(payload) and uses_default_closure_date(payload, str(closure.get("reviewed_on", ""))):
        st.session_state[reviewed_on_key] = default_review_date
    closure["reviewed_on"] = review_cols[1].date_input(
        "Fecha de revision",
        key=reviewed_on_key,
        disabled=not allow_closure_edits,
    ).isoformat()


def validate_payload(payload: dict[str, Any]) -> list[str]:
    errors: list[str] = []
    definition = get_form_definition(payload["metadata"]["form_key"])
    for day in range(1, 32):
        record = payload["daily_records"][str(day)]
        if not record["active"]:
            continue
        active_slots = active_slot_indices(record)
        if not active_slots:
            continue
        for metric in definition["metrics"]:
            metric_values = record.get(metric["key"], ["", "", ""])
            if not all(str(metric_values[index] if index < len(metric_values) else "").strip() for index in active_slots):
                metric_label = get_primary_metric_display_label(payload, metric).lower()
                errors.append(f"Dia {day}: faltan capturas de {metric_label}.")
        performed_values = record.get("performed_by_slots", ["", "", ""])
        if not all(str(performed_values[index] if index < len(performed_values) else "").strip() for index in active_slots):
            errors.append(f"Dia {day}: faltan responsables en una o mas horas.")
        if not record["verified_by"].strip():
            errors.append(f"Dia {day}: falta 'Verifico'.")
        if not record["recorded_on"].strip():
            errors.append(f"Dia {day}: falta la fecha.")

    if not payload["monthly_closure"]["reviewed_by"].strip():
        errors.append("Falta capturar quien reviso el formato.")

    return errors


def parse_optional_iso_date(value: str) -> date | None:
    text = str(value or "").strip()
    if not text:
        return None
    try:
        return datetime.fromisoformat(text).date()
    except ValueError:
        return None


def format_traceability_date(value: str) -> str:
    parsed = parse_optional_iso_date(value)
    return parsed.strftime("%d/%m/%Y") if parsed else "-"


def build_history_dataframe(payload: dict[str, Any], days_back: int) -> pd.DataFrame:
    form_key = str(payload["metadata"]["form_key"])
    equipment_code = str(payload["metadata"]["equipment_code"])
    periods = list_periods_for_equipment(form_key, equipment_code)
    if not periods:
        return pd.DataFrame()

    definition = get_form_definition(form_key)
    primary_metric = definition["metrics"][0]
    secondary_metric = definition["metrics"][1] if len(definition["metrics"]) > 1 else None
    cutoff = date.today() - timedelta(days=days_back - 1)
    rows: list[dict[str, Any]] = []

    for period in periods:
        period_payload = load_saved_payload(
            form_key=form_key,
            equipment_code=equipment_code,
            year=int(period["year"]),
            month=int(period["month"]),
        )
        for day in range(1, 32):
            record = period_payload["daily_records"][str(day)]
            if not record["active"]:
                continue
            active_slots = active_slot_indices(record)
            if not active_slots:
                continue

            recorded_date = parse_optional_iso_date(record.get("recorded_on", "")) or get_row_period_date(period_payload, day)
            if recorded_date < cutoff or recorded_date > date.today():
                continue

            primary_values = (
                record["corrected_temperatures"]
                if any(record["corrected_temperatures"]) and primary_metric.get("corrected")
                else record["measured_temperatures"]
            )
            primary_numeric = [parse_measurement_number(value) for value in primary_values]
            primary_numeric = [
                numeric_value if index in active_slots else None
                for index, numeric_value in enumerate(primary_numeric)
            ]
            if all(value is None for value in primary_numeric):
                continue

            row: dict[str, Any] = {
                "fecha": recorded_date,
                "dia": day,
            }
            valid_primary = [float(value) for value in primary_numeric if value is not None]
            row["promedio_primario"] = sum(valid_primary) / len(valid_primary)
            row["minimo_primario"] = min(valid_primary)
            row["maximo_primario"] = max(valid_primary)
            for slot_label, numeric_value in zip(TIME_SLOTS, primary_numeric):
                row[slot_label] = numeric_value

            if secondary_metric is not None:
                secondary_numeric = [parse_measurement_number(value) for value in record[secondary_metric["key"]]]
                secondary_numeric = [
                    numeric_value if index in active_slots else None
                    for index, numeric_value in enumerate(secondary_numeric)
                ]
                valid_secondary = [float(value) for value in secondary_numeric if value is not None]
                row["promedio_secundario"] = (
                    sum(valid_secondary) / len(valid_secondary) if valid_secondary else None
                )
            rows.append(row)

    if not rows:
        return pd.DataFrame()

    dataframe = pd.DataFrame(rows).sort_values("fecha")
    dataframe["fecha"] = pd.to_datetime(dataframe["fecha"])
    return dataframe


MASTER_COLUMN_ALIASES = {
    "tipo": ["TIPO DE DOCUMENTO", "TIPO"],
    "nombre": ["NOMBRE DEL DOCUMENTO", "DOCUMENTO", "NOMBRE"],
    "codigo": ["CODIGO LIT", "CODIGO", "CÓDIGO LIT"],
    "fecha_alta": ["FECHA DE ALTA", "ALTA"],
    "fecha_vigencia": ["FECHA DE VIGENCIA", "VIGENCIA"],
    "almacenamiento": ["ALMACENAMIENTO", "UBICACION", "UBICACIÓN"],
}


def normalize_master_column_name(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(character for character in text if not unicodedata.combining(character))
    return re.sub(r"\s+", " ", text).strip().upper()


def find_master_column(columns: list[str], alias_key: str) -> str | None:
    candidates = {
        normalize_master_column_name(candidate)
        for candidate in MASTER_COLUMN_ALIASES.get(alias_key, [])
    }
    for column in columns:
        if normalize_master_column_name(column) in candidates:
            return column
    return None


def parse_master_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and not pd.isna(value):
        return date(1899, 12, 30) + timedelta(days=int(value))

    text = str(value).strip()
    if not text or normalize_master_column_name(text) in {"PENDIENTE", "N/A", "NA"}:
        return None
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).date()
    except ValueError:
        pass
    for date_format in ("%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%m-%d-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, date_format).date()
        except ValueError:
            continue
    return None


def normalize_master_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value).strip()


@st.cache_data(show_spinner=False, ttl=300)
def load_master_table_cached() -> dict[str, Any]:
    return get_master_table()


def build_master_dataframe(table: dict[str, Any]) -> pd.DataFrame:
    columns = [str(column) for column in table.get("columns", [])]
    rows: list[dict[str, Any]] = []
    for row in table.get("rows", []):
        values = list(row.get("values", []))
        if len(values) < len(columns):
            values.extend([""] * (len(columns) - len(values)))
        values = values[: len(columns)]
        if not any(normalize_master_cell(value) for value in values):
            continue
        row_data = dict(zip(columns, values))
        row_data["_row_index"] = int(row.get("index", len(rows)))
        rows.append(row_data)
    return pd.DataFrame(rows, columns=[*columns, "_row_index"])


def build_master_alerts(dataframe: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    if dataframe.empty:
        return pd.DataFrame()
    vigencia_col = find_master_column(columns, "fecha_vigencia")
    if not vigencia_col or vigencia_col not in dataframe.columns:
        return pd.DataFrame()

    codigo_col = find_master_column(columns, "codigo")
    nombre_col = find_master_column(columns, "nombre")
    tipo_col = find_master_column(columns, "tipo")
    today = date.today()
    alert_rows = []
    for _, row in dataframe.iterrows():
        vigencia = parse_master_date(row.get(vigencia_col))
        days_left = (vigencia - today).days if vigencia else None
        if days_left is None:
            status = "Sin fecha"
        elif days_left < 0:
            status = "Vencido"
        elif days_left <= 30:
            status = "Vence en 30 dias"
        elif days_left <= 60:
            status = "Vence en 60 dias"
        elif days_left <= 90:
            status = "Vence en 90 dias"
        else:
            status = "Vigente"
        alert_rows.append(
            {
                "Estado": status,
                "Dias restantes": "" if days_left is None else days_left,
                "Fecha de vigencia": vigencia.strftime("%d/%m/%Y") if vigencia else "",
                "Codigo": row.get(codigo_col, "") if codigo_col else "",
                "Documento": row.get(nombre_col, "") if nombre_col else "",
                "Tipo": row.get(tipo_col, "") if tipo_col else "",
            }
        )

    alerts_df = pd.DataFrame(alert_rows)
    severity_order = {
        "Vencido": 0,
        "Vence en 30 dias": 1,
        "Vence en 60 dias": 2,
        "Vence en 90 dias": 3,
        "Sin fecha": 4,
        "Vigente": 5,
    }
    alerts_df["_order"] = alerts_df["Estado"].map(severity_order).fillna(9)
    return alerts_df.sort_values(["_order", "Dias restantes"], kind="stable").drop(columns=["_order"])


def render_master_list() -> None:
    st.subheader("Lista maestra")
    st.caption("Consulta vigencias desde el Excel de OneDrive y edita la tabla central sin duplicar archivos.")

    if not microsoft_graph_enabled():
        st.warning(
            "Falta configurar Microsoft Graph en secrets: microsoft_tenant_id, "
            "microsoft_client_id, microsoft_client_secret, microsoft_shared_url y microsoft_table_name."
        )
        return

    refresh_col, link_col = st.columns([1, 3])
    if refresh_col.button("Actualizar lista", use_container_width=True):
        load_master_table_cached.clear()
        st.rerun()

    try:
        table = load_master_table_cached()
    except MicrosoftGraphError as exc:
        st.error(f"No se pudo cargar la lista maestra: {exc}")
        return

    columns = [str(column) for column in table.get("columns", [])]
    dataframe = build_master_dataframe(table)
    web_url = str(table.get("web_url", ""))
    if web_url:
        link_col.link_button("Abrir en OneDrive", web_url, use_container_width=True)

    if dataframe.empty:
        st.info("La tabla de OneDrive no tiene registros utiles para mostrar.")
        return

    alerts_df = build_master_alerts(dataframe, columns)
    if not alerts_df.empty:
        alert_filter = st.selectbox(
            "Filtro de alertas",
            options=["Criticos", "90 dias", "Todos"],
            key="master_alert_filter",
        )
        if alert_filter == "Criticos":
            display_alerts = alerts_df[
                alerts_df["Estado"].isin(["Vencido", "Vence en 30 dias", "Sin fecha"])
            ]
        elif alert_filter == "90 dias":
            display_alerts = alerts_df[
                alerts_df["Estado"].isin(
                    ["Vencido", "Vence en 30 dias", "Vence en 60 dias", "Vence en 90 dias", "Sin fecha"]
                )
            ]
        else:
            display_alerts = alerts_df

        st.markdown("**Alertas de vigencia**")
        st.dataframe(display_alerts, use_container_width=True, hide_index=True)
    else:
        st.info("No encontre una columna de Fecha de vigencia para generar alertas.")

    st.markdown("**Registros de la lista maestra**")
    editable = can_edit_master_list()
    if editable:
        edited_df = st.data_editor(
            dataframe,
            use_container_width=True,
            hide_index=True,
            num_rows="fixed",
            disabled=["_row_index"],
            column_config={"_row_index": None},
            key="master_list_editor",
        )
        save_col, _ = st.columns([1, 3])
        if save_col.button("Guardar cambios en OneDrive", use_container_width=True):
            changes_count = 0
            code_col = find_master_column(columns, "codigo")
            for row_position in range(len(dataframe)):
                original_row = dataframe.iloc[row_position]
                edited_row = edited_df.iloc[row_position]
                changed_columns = [
                    column
                    for column in columns
                    if normalize_master_cell(original_row.get(column)) != normalize_master_cell(edited_row.get(column))
                ]
                if not changed_columns:
                    continue

                row_values = [
                    "" if pd.isna(edited_row.get(column)) else edited_row.get(column)
                    for column in columns
                ]
                update_master_table_row(int(original_row["_row_index"]), row_values)
                changes_count += len(changed_columns)
                code_value = normalize_master_cell(edited_row.get(code_col)) if code_col else f"fila {row_position + 1}"
                for column in changed_columns:
                    log_activity(
                        "actualizar_lista_maestra",
                        f"{code_value}: {column}: {normalize_master_cell(original_row.get(column))} -> {normalize_master_cell(edited_row.get(column))}",
                        {"metadata": {}},
                    )

            if changes_count:
                load_master_table_cached.clear()
                st.success(f"Se guardaron {changes_count} cambio(s) en OneDrive.")
                st.rerun()
            else:
                st.info("No hubo cambios para guardar.")

        with st.expander("Agregar documento"):
            with st.form("master_list_add_row"):
                new_values_by_column: dict[str, Any] = {}
                input_columns = st.columns(3)
                for index, column in enumerate(columns):
                    new_values_by_column[column] = input_columns[index % 3].text_input(column, key=f"new_master_{column}")
                if st.form_submit_button("Agregar a OneDrive", use_container_width=True):
                    new_values = [new_values_by_column[column] for column in columns]
                    if not any(normalize_master_cell(value) for value in new_values):
                        st.warning("Captura al menos un dato antes de agregar la fila.")
                    else:
                        add_master_table_row(new_values)
                        log_activity("agregar_lista_maestra", "Agrego documento a lista maestra", {"metadata": {}})
                        load_master_table_cached.clear()
                        st.success("Documento agregado a OneDrive.")
                        st.rerun()
    else:
        st.dataframe(dataframe.drop(columns=["_row_index"], errors="ignore"), use_container_width=True, hide_index=True)
        st.caption("Solo admin puede editar la lista maestra desde la app.")


def render_reports(payload: dict[str, Any]) -> None:
    st.subheader("Reportes")
    st.caption("Consulta el comportamiento historico del equipo por ventana semanal, mensual y semestral.")
    report_ranges = {
        "Semanal": 7,
        "1 mes": 30,
        "3 meses": 90,
        "6 meses": 180,
    }
    selected_range_label = st.selectbox(
        "Ventana del reporte",
        options=list(report_ranges.keys()),
        key=f"report_range_{get_period_key(payload)}",
    )
    history_df = build_history_dataframe(payload, report_ranges[selected_range_label])
    if history_df.empty:
        st.info("Aun no hay suficientes registros historicos para este equipo en la ventana seleccionada.")
        return

    primary_metric = get_form_definition(payload["metadata"]["form_key"])["metrics"][0]
    primary_label = get_primary_metric_display_label(payload, primary_metric)
    summary_cols = st.columns(4)
    summary_cols[0].metric("Dias con captura", f"{len(history_df)}")
    summary_cols[1].metric("Promedio", format_decimal_value(float(history_df["promedio_primario"].mean())))
    summary_cols[2].metric("Minimo", format_decimal_value(float(history_df["minimo_primario"].min())))
    summary_cols[3].metric("Maximo", format_decimal_value(float(history_df["maximo_primario"].max())))

    st.markdown(f"**{primary_label}: promedio, minimo y maximo diarios**")
    st.line_chart(
        history_df.set_index("fecha")[["promedio_primario", "minimo_primario", "maximo_primario"]],
        use_container_width=True,
    )

    st.markdown(f"**{primary_label}: comportamiento por horario**")
    slot_columns = [slot for slot in TIME_SLOTS if slot in history_df.columns]
    st.line_chart(history_df.set_index("fecha")[slot_columns], use_container_width=True)

    if "promedio_secundario" in history_df.columns and history_df["promedio_secundario"].notna().any():
        secondary_label = "%CO2" if payload["metadata"]["form_key"] == "incubadoras" else "Variable secundaria"
        st.markdown(f"**{secondary_label}: promedio diario**")
        st.line_chart(history_df.set_index("fecha")[["promedio_secundario"]], use_container_width=True)

    display_df = history_df.copy()
    display_df["fecha"] = display_df["fecha"].dt.strftime("%Y-%m-%d")
    for column in [*TIME_SLOTS, "promedio_primario", "minimo_primario", "maximo_primario", "promedio_secundario"]:
        if column in display_df.columns:
            display_df[column] = display_df[column].apply(
                lambda value: "" if pd.isna(value) else format_decimal_value(float(value))
            )
    st.dataframe(display_df, use_container_width=True, hide_index=True)
    st.download_button(
        "Descargar reporte CSV",
        data=display_df.to_csv(index=False).encode("utf-8-sig"),
        file_name=f"reporte_{payload['metadata']['form_key']}_{payload['metadata']['equipment_code']}_{selected_range_label.replace(' ', '_').lower()}.csv",
        mime="text/csv",
        use_container_width=True,
    )


def render_traceability_and_validation(payload: dict[str, Any]) -> None:
    st.subheader("Trazabilidad y validacion")
    st.caption("Programa y consulta mantenimiento, calibracion, calificacion y validacion de la aplicacion por equipo.")
    validation_errors = validate_payload(payload)
    if validation_errors:
        st.warning("La validacion automatica del periodo actual tiene pendientes.")
        for error in validation_errors:
            st.write(f"- {error}")
    else:
        st.success("La validacion automatica del periodo actual no detecto pendientes.")

    form_key = str(payload["metadata"]["form_key"])
    equipment_code = str(payload["metadata"]["equipment_code"])
    entries = list_traceability_entries(form_key, equipment_code)
    upcoming_entries = [
        entry
        for entry in entries
        if entry.get("status") == "programado"
        and (parse_optional_iso_date(entry.get("scheduled_for", "")) or date.max) <= date.today() + timedelta(days=30)
    ]
    if upcoming_entries:
        st.info(f"Hay {len(upcoming_entries)} evento(s) programado(s) dentro de los proximos 30 dias.")

    summary_rows = [
        {
            "Tipo": TRACEABILITY_TYPES.get(entry["entry_type"], entry["entry_type"]),
            "Estado": TRACEABILITY_STATUSES.get(entry["status"], entry["status"]),
            "Programado": format_traceability_date(entry.get("scheduled_for", "")),
            "Realizado": format_traceability_date(entry.get("completed_on", "")),
            "Proveedor / responsable": entry.get("provider", ""),
            "Notas": entry.get("notes", ""),
        }
        for entry in entries
    ]
    if summary_rows:
        st.dataframe(pd.DataFrame(summary_rows), use_container_width=True, hide_index=True)
    else:
        st.caption("Todavia no hay eventos programados para este equipo.")

    traceability_type_options = list(TRACEABILITY_TYPES.keys())
    traceability_status_options = list(TRACEABILITY_STATUSES.keys())

    if can_manage_traceability():
        with st.form(f"traceability_new_{get_period_key(payload)}"):
            st.markdown("**Registrar o programar evento**")
            type_col, status_col = st.columns(2)
            entry_type = type_col.selectbox(
                "Tipo",
                options=traceability_type_options,
                format_func=lambda value: TRACEABILITY_TYPES[value],
            )
            status = status_col.selectbox(
                "Estado",
                options=traceability_status_options,
                format_func=lambda value: TRACEABILITY_STATUSES[value],
            )
            schedule_col, completed_col = st.columns(2)
            scheduled_for = schedule_col.date_input("Fecha programada", value=date.today())
            completed_on = completed_col.date_input("Fecha realizada", value=date.today())
            provider = st.text_input("Proveedor o responsable")
            notes = st.text_area("Notas", height=90)
            if st.form_submit_button("Guardar evento", use_container_width=True):
                saved_entry = save_traceability_entry(
                    {
                        "form_key": form_key,
                        "equipment_code": equipment_code,
                        "entry_type": entry_type,
                        "status": status,
                        "scheduled_for": scheduled_for.isoformat(),
                        "completed_on": completed_on.isoformat() if status == "realizado" else "",
                        "provider": provider,
                        "notes": notes,
                    }
                )
                log_activity(
                    "guardar_trazabilidad",
                    f"{TRACEABILITY_TYPES.get(saved_entry['entry_type'], saved_entry['entry_type'])} -> {TRACEABILITY_STATUSES.get(saved_entry['status'], saved_entry['status'])}",
                    payload,
                )
                st.success("Evento guardado.")
                st.rerun()

        for entry in entries:
            title = f"{TRACEABILITY_TYPES.get(entry['entry_type'], entry['entry_type'])} | {TRACEABILITY_STATUSES.get(entry['status'], entry['status'])} | {format_traceability_date(entry.get('scheduled_for', ''))}"
            with st.expander(title):
                entry_type_key = f"traceability_type_{entry['id']}"
                entry_status_key = f"traceability_status_{entry['id']}"
                scheduled_key = f"traceability_scheduled_{entry['id']}"
                completed_key = f"traceability_completed_{entry['id']}"
                provider_key = f"traceability_provider_{entry['id']}"
                notes_key = f"traceability_notes_{entry['id']}"

                entry_type_value = st.selectbox(
                    "Tipo",
                    options=traceability_type_options,
                    index=traceability_type_options.index(entry["entry_type"]) if entry["entry_type"] in traceability_type_options else 0,
                    format_func=lambda value: TRACEABILITY_TYPES[value],
                    key=entry_type_key,
                )
                status_value = st.selectbox(
                    "Estado",
                    options=traceability_status_options,
                    index=traceability_status_options.index(entry["status"]) if entry["status"] in traceability_status_options else 0,
                    format_func=lambda value: TRACEABILITY_STATUSES[value],
                    key=entry_status_key,
                )
                scheduled_value = st.date_input(
                    "Fecha programada",
                    value=parse_optional_iso_date(entry.get("scheduled_for", "")) or date.today(),
                    key=scheduled_key,
                )
                completed_value = st.date_input(
                    "Fecha realizada",
                    value=parse_optional_iso_date(entry.get("completed_on", "")) or date.today(),
                    key=completed_key,
                )
                provider_value = st.text_input(
                    "Proveedor o responsable",
                    value=str(entry.get("provider", "")),
                    key=provider_key,
                )
                notes_value = st.text_area(
                    "Notas",
                    value=str(entry.get("notes", "")),
                    key=notes_key,
                    height=90,
                )
                action_cols = st.columns(2)
                if action_cols[0].button("Actualizar", key=f"traceability_save_{entry['id']}", use_container_width=True):
                    save_traceability_entry(
                        {
                            "id": entry["id"],
                            "form_key": form_key,
                            "equipment_code": equipment_code,
                            "entry_type": entry_type_value,
                            "status": status_value,
                            "scheduled_for": scheduled_value.isoformat(),
                            "completed_on": completed_value.isoformat() if status_value == "realizado" else "",
                            "provider": provider_value,
                            "notes": notes_value,
                            "created_by": entry.get("created_by", ""),
                        }
                    )
                    log_activity("actualizar_trazabilidad", TRACEABILITY_TYPES.get(entry_type_value, entry_type_value), payload)
                    st.rerun()
                if action_cols[1].button("Eliminar", key=f"traceability_delete_{entry['id']}", use_container_width=True):
                    delete_traceability_entry(entry["id"], form_key, equipment_code)
                    log_activity("eliminar_trazabilidad", TRACEABILITY_TYPES.get(entry["entry_type"], entry["entry_type"]), payload)
                    st.rerun()
    else:
        st.caption("Solo admin puede programar o actualizar esta agenda.")


def populate_template(payload: dict[str, Any]) -> BytesIO:
    form_key = payload["metadata"]["form_key"]
    definition = get_form_definition(form_key)
    workbook = load_template_workbook(form_key)
    target_sheet_name = str(payload["metadata"]["equipment_code"])
    worksheet = workbook[target_sheet_name]
    ensure_status_row_merges(worksheet, payload)

    metadata = payload["metadata"]
    metadata_cells = payload.get("metadata_cells", {})
    editable_metadata_map = {
        "equipment_name": metadata.get("equipment_name", ""),
        "brand": metadata.get("brand", ""),
        "model": metadata.get("model", ""),
        "inventory_code": metadata.get("inventory_code", ""),
        "serial_number": metadata.get("serial_number", ""),
        "temperature_label": metadata.get("temperature_label", ""),
        "minimum_label": metadata.get("minimum_label", ""),
        "maximum_label": metadata.get("maximum_label", ""),
        "secondary_label": metadata.get("secondary_label", ""),
        "secondary_minimum_label": metadata.get("secondary_minimum_label", ""),
        "secondary_maximum_label": metadata.get("secondary_maximum_label", ""),
    }
    for field_key, value in editable_metadata_map.items():
        cell = metadata_cells.get(field_key)
        if cell:
            write_template_cell(worksheet, cell, value)

    month_cell = HEADER_CELL_MAP["month"]
    year_cell = HEADER_CELL_MAP["year"]
    if form_key == "condiciones_ambientales":
        month_cell = "V6"
        year_cell = "AO6"
    write_template_cell(worksheet, month_cell, MONTHS[payload["metadata"]["month"]])
    write_template_cell(worksheet, year_cell, payload["metadata"]["year"])

    if definition["supports_corrections"]:
        range_cells = payload.get("correction_range_cells", {})
        for factor_key, cell in range_cells.items():
            band = payload.get("correction_bands", {}).get(factor_key)
            if not cell or not isinstance(band, dict):
                continue
            label = str(band.get("label", "")).strip()
            if not label:
                separator = str(band.get("separator", "-")).strip() or "-"
                label = build_range_label(
                    float(band.get("min", 0)),
                    float(band.get("max", 0)),
                    separator,
                )
            write_template_cell(worksheet, cell, label)
        factor_cells = payload.get("correction_cells", {})
        for factor_key, cell in factor_cells.items():
            if factor_key not in payload["correction_factors"]:
                continue
            factor_value = float(payload["correction_factors"][factor_key])
            if payload["correction_operations"][factor_key] == "-":
                factor_value *= -1
            write_template_cell(worksheet, cell, factor_value)

    footer_map = definition["layout"]["footer"]

    for day in range(1, 32):
        record = payload["daily_records"][str(day)]
        row_group = get_row_group(payload, day)
        start_col = get_day_column(day, 0)

        if not record["active"]:
            for index in range(3):
                write_slot_value(
                    worksheet,
                    row_group["metric_1"],
                    start_col + index,
                    "N/A",
                    font_size=18,
                    rotate_like_hours=True,
                )
                if "metric_2" in row_group:
                    write_slot_value(
                        worksheet,
                        row_group["metric_2"],
                        start_col + index,
                        "N/A",
                        font_size=18,
                        rotate_like_hours=True,
                    )
                write_slot_value(
                    worksheet,
                    row_group["performed_by"],
                    start_col + index,
                    "NO LABORADO",
                    font_size=18,
                    rotate_like_hours=True,
                )
            write_day_status(worksheet, row_group["verified_by"], start_col, "N/A")
            write_day_status(
                worksheet,
                row_group["date"],
                start_col,
                "N/A",
            )
            continue

        primary_values = (
            record["corrected_temperatures"]
            if any(record["corrected_temperatures"]) and definition["metrics"][0].get("corrected")
            else record["measured_temperatures"]
        )
        primary_metric = definition["metrics"][0]
        primary_unit = primary_metric["unit"]
        for index, temperature in enumerate(primary_values):
            if is_slot_canceled(record, index):
                write_slot_value(
                    worksheet,
                    row_group["metric_1"],
                    start_col + index,
                    "N/A",
                    font_size=18,
                    rotate_like_hours=True,
                )
                continue
            formatted_value = format_metric_value(payload, primary_metric, temperature)
            write_slot_value(
                worksheet,
                row_group["metric_1"],
                start_col + index,
                f"{formatted_value} {primary_unit}".strip() if formatted_value else "",
                font_size=18,
                rotate_like_hours=True,
            )

        if len(definition["metrics"]) > 1 and "metric_2" in row_group:
            second_metric = definition["metrics"][1]
            for index, value in enumerate(record[second_metric["key"]]):
                if is_slot_canceled(record, index):
                    write_slot_value(
                        worksheet,
                        row_group["metric_2"],
                        start_col + index,
                        "N/A",
                        font_size=18,
                        rotate_like_hours=True,
                    )
                    continue
                formatted_second_value = format_metric_value(payload, second_metric, value)
                write_slot_value(
                    worksheet,
                    row_group["metric_2"],
                    start_col + index,
                    f"{formatted_second_value} {second_metric['unit']}".strip() if formatted_second_value else "",
                    font_size=18,
                    rotate_like_hours=True,
                )

        for index, performed_by in enumerate(record["performed_by_slots"]):
            if is_slot_canceled(record, index):
                write_slot_value(
                    worksheet,
                    row_group["performed_by"],
                    start_col + index,
                    "CANCELADO",
                    font_size=18,
                    rotate_like_hours=True,
                )
                continue
            write_signature_or_text_slot(
                worksheet,
                row_group["performed_by"],
                start_col + index,
                performed_by,
                width=70,
                height=170,
            )
        if is_day_fully_canceled(record):
            write_day_status(worksheet, row_group["verified_by"], start_col, "N/A")
            write_day_status(worksheet, row_group["date"], start_col, "N/A")
            continue
        write_signature_or_text_status(
            worksheet,
            row_group["verified_by"],
            start_col,
            record["verified_by"],
            width=120,
            height=42,
        )
        write_day_status(
            worksheet,
            row_group["date"],
            start_col,
            get_effective_record_date(record),
        )

    write_observations_cell(worksheet, footer_map["observations"], compose_observations_export_text(payload))
    write_signature_or_text_cell(
        worksheet,
        footer_map["reviewed_by"],
        payload["monthly_closure"]["reviewed_by"],
        width=180,
        height=52,
    )
    write_template_cell(
        worksheet,
        footer_map["reviewed_on"],
        normalize_excel_date(payload["monthly_closure"]["reviewed_on"]),
    )

    for index, sheet_name in enumerate(workbook.sheetnames):
        sheet = workbook[sheet_name]
        if sheet_name == target_sheet_name:
            sheet.sheet_state = "visible"
            workbook.active = index
        else:
            sheet.sheet_state = "hidden"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def load_template_workbook(form_key: str, data_only: bool = False):
    return load_workbook(BytesIO(get_template_bytes(form_key)), data_only=data_only)


def resolve_writable_coordinate(worksheet: Any, coordinate: str) -> str:
    for merged_range in worksheet.merged_cells.ranges:
        if coordinate in merged_range:
            return worksheet.cell(merged_range.min_row, merged_range.min_col).coordinate
    return coordinate


def write_template_cell(worksheet: Any, coordinate: str, value: Any) -> None:
    worksheet[resolve_writable_coordinate(worksheet, coordinate)] = value


def write_observations_cell(worksheet: Any, coordinate: str, value: str) -> None:
    writable_coordinate = resolve_writable_coordinate(worksheet, coordinate)
    cell = worksheet[writable_coordinate]
    cell.value = value
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    merged_range = get_merged_range_for_coordinate(worksheet, writable_coordinate)
    if merged_range is None:
        min_col = max_col = cell.column
        min_row = max_row = cell.row
    else:
        min_col = merged_range.min_col
        max_col = merged_range.max_col
        min_row = merged_range.min_row
        max_row = merged_range.max_row

    total_width_pixels = 0.0
    for column in range(min_col, max_col + 1):
        total_width_pixels += column_width_to_pixels(
            worksheet.column_dimensions[get_column_letter(column)].width
        )

    approx_chars_per_line = max(int(total_width_pixels / 18), 45)
    content_lines = str(value or "").splitlines() or [""]
    estimated_lines = 0
    for line in content_lines:
        width_based_lines = max(1, (len(line) + approx_chars_per_line - 1) // approx_chars_per_line)
        conservative_lines = max(1, (len(line) + 79) // 80)
        estimated_lines += max(width_based_lines, conservative_lines)

    total_rows = max_row - min_row + 1
    total_height_points = max(60.0, estimated_lines * 18.0)
    row_height_points = total_height_points / total_rows
    for row in range(min_row, max_row + 1):
        current_height = worksheet.row_dimensions[row].height
        worksheet.row_dimensions[row].height = max(current_height or 0, row_height_points)


def normalize_signature_text(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_text = "".join(char for char in normalized if not unicodedata.combining(char))
    ascii_text = re.sub(r"[_\-.,/]+", " ", ascii_text.lower())
    ascii_text = re.sub(r"\s+", " ", ascii_text).strip()
    return ascii_text


def strip_signature_suffix(file_stem: str) -> str:
    cleaned = re.sub(r"_(azul|negro|firma)$", "", file_stem, flags=re.IGNORECASE)
    cleaned = re.sub(r"\s+\d+$", "", cleaned).strip()
    return cleaned


def build_signature_aliases(tokens: list[str], normalized_name: str) -> dict[str, int]:
    aliases: dict[str, int] = {normalized_name: 1000}
    collapsed_name = normalized_name.replace(" ", "")
    if collapsed_name:
        aliases[collapsed_name] = 980

    for token in tokens:
        aliases.setdefault(token, 120)
        aliases.setdefault(f"{token[0]} {token}", 890)
        aliases.setdefault(f"{token[0]}.{token}", 890)

    for index, token in enumerate(tokens):
        for other_index, other_token in enumerate(tokens):
            if index == other_index:
                continue

            aliases.setdefault(f"{token} {other_token}", 340)
            aliases.setdefault(f"{other_token} {token}", 330)

            initial = other_token[0]
            aliases.setdefault(f"{initial} {token}", 900)
            aliases.setdefault(f"{initial}.{token}", 900)
            aliases.setdefault(f"{token} {initial}", 860)
            aliases.setdefault(f"{token}.{initial}", 860)

            aliases.setdefault(f"{other_token} {token}", 780)
            aliases.setdefault(f"{token} {other_token}", 780)

    if len(tokens) >= 2:
        primary_surname = tokens[0]
        for given_token in tokens[1:]:
            initial = given_token[0]
            aliases[f"{initial} {primary_surname}"] = max(aliases.get(f"{initial} {primary_surname}", 0), 950)
            aliases[f"{initial}.{primary_surname}"] = max(aliases.get(f"{initial}.{primary_surname}", 0), 950)
            aliases[f"{given_token} {primary_surname}"] = max(aliases.get(f"{given_token} {primary_surname}", 0), 920)

    return aliases


def get_local_signature_cache_key() -> tuple[str, ...]:
    if not SIGNATURES_DIR.exists():
        return ()
    return tuple(sorted(path.name.lower() for path in SIGNATURES_DIR.glob("*.png")))


@st.cache_data(show_spinner=False, ttl=300)
def _load_signature_catalog_cached(
    _storage_cache_key: tuple[bool, str, str],
    _local_cache_key: tuple[str, ...],
) -> list[dict[str, Any]]:
    catalog: list[dict[str, Any]] = []
    remote_assets: list[dict[str, str]] = []
    if signatures_storage_enabled():
        try:
            remote_assets = sorted(list_remote_signature_assets(), key=lambda asset: asset["name"].lower())
        except Exception:
            remote_assets = []

    for asset in remote_assets:
        asset_name = asset["name"]
        display_name = strip_signature_suffix(Path(asset_name).stem)
        normalized_name = normalize_signature_text(display_name)
        tokens = normalized_name.split()
        catalog.append(
            {
                "asset_name": asset_name,
                "storage_path": asset["path"],
                "local_path": None,
                "display_name": display_name,
                "normalized_name": normalized_name,
                "tokens": tokens,
                "aliases": build_signature_aliases(tokens, normalized_name),
            }
        )

    if SIGNATURES_DIR.exists():
        existing_names = {candidate["asset_name"].lower() for candidate in catalog}
        for path in sorted(SIGNATURES_DIR.glob("*.png")):
            if path.name.lower() in existing_names:
                continue
            display_name = strip_signature_suffix(path.stem)
            normalized_name = normalize_signature_text(display_name)
            tokens = normalized_name.split()
            catalog.append(
                {
                    "asset_name": path.name,
                    "storage_path": None,
                    "local_path": path,
                    "display_name": display_name,
                    "normalized_name": normalized_name,
                    "tokens": tokens,
                    "aliases": build_signature_aliases(tokens, normalized_name),
                }
            )
    return catalog


def load_signature_catalog() -> list[dict[str, Any]]:
    return _load_signature_catalog_cached(
        get_signatures_storage_cache_key(),
        get_local_signature_cache_key(),
    )


def find_signature_candidate(person_name: str) -> dict[str, Any] | None:
    normalized_input = normalize_signature_text(person_name)
    if not normalized_input:
        return None

    catalog = load_signature_catalog()
    if not catalog:
        return None

    raw_input = unicodedata.normalize("NFKD", person_name)
    ascii_input = "".join(char for char in raw_input if not unicodedata.combining(char))
    dotted_match = re.search(r"\.\s*([A-Za-z]+)", ascii_input)
    if dotted_match:
        surname_after_dot = normalize_signature_text(dotted_match.group(1))
        surname_matches = [
            candidate
            for candidate in catalog
            if surname_after_dot and surname_after_dot in candidate["tokens"]
        ]
        unique_surname_matches = {candidate["asset_name"].lower(): candidate for candidate in surname_matches}
        if len(unique_surname_matches) == 1:
            return next(iter(unique_surname_matches.values()))

    best_score = -1
    best_matches: list[dict[str, Any]] = []
    collapsed_input = normalized_input.replace(" ", "")
    input_tokens = normalized_input.split()

    for candidate in catalog:
        score = 0
        aliases = candidate["aliases"]
        if normalized_input in aliases:
            score = max(score, aliases[normalized_input])
        if collapsed_input in aliases:
            score = max(score, aliases[collapsed_input])
        if collapsed_input and collapsed_input == candidate["normalized_name"].replace(" ", ""):
            score = max(score, 970)
        if input_tokens and all(token in candidate["tokens"] for token in input_tokens):
            score = max(score, 400 + (len(input_tokens) * 10))
        if len(input_tokens) == 1 and input_tokens[0] in candidate["tokens"]:
            score = max(score, 180)

        if score > best_score:
            best_score = score
            best_matches = [candidate]
        elif score > 0 and score == best_score:
            best_matches.append(candidate)

    unique_matches = {candidate["asset_name"].lower(): candidate for candidate in best_matches}
    if best_score <= 0 or len(unique_matches) != 1:
        return None
    return next(iter(unique_matches.values()))


def get_signature_display_name(person_name: str) -> str | None:
    signature_candidate = find_signature_candidate(person_name)
    if signature_candidate is None:
        return None
    return str(signature_candidate["display_name"])


def get_signature_image_source(signature_candidate: dict[str, Any]) -> BytesIO | Path | None:
    storage_path = signature_candidate.get("storage_path")
    if storage_path:
        try:
            signature_bytes = download_signature_bytes(str(storage_path))
            image_buffer = BytesIO(signature_bytes)
            image_buffer.seek(0)
            return image_buffer
        except Exception:
            pass

    local_path = signature_candidate.get("local_path")
    if isinstance(local_path, Path) and local_path.exists():
        return local_path
    return None


def add_signature_image(
    worksheet: Any,
    coordinate: str,
    image_source: BytesIO | Path,
    width: int,
    height: int,
    rotate_vertical: bool = False,
) -> None:
    writable_coordinate = resolve_writable_coordinate(worksheet, coordinate)
    worksheet[writable_coordinate] = None
    image_buffer = prepare_signature_image_buffer(image_source, rotate_vertical=rotate_vertical)
    image = XLImage(image_buffer)
    fitted_width, fitted_height = fit_signature_dimensions(
        worksheet,
        writable_coordinate,
        original_width=image.width,
        original_height=image.height,
        requested_width=width,
        requested_height=height,
    )
    image.width = fitted_width
    image.height = fitted_height
    image.anchor = build_signature_anchor(worksheet, writable_coordinate, fitted_width, fitted_height)
    worksheet.add_image(image)


def prepare_signature_image_buffer(image_source: BytesIO | Path, rotate_vertical: bool = False) -> BytesIO:
    if isinstance(image_source, Path):
        image = PILImage.open(image_source).convert("RGBA")
    else:
        image_source.seek(0)
        image = PILImage.open(image_source).convert("RGBA")
    alpha_bbox = image.getchannel("A").getbbox() if "A" in image.getbands() else None
    if alpha_bbox is not None:
        image = image.crop(alpha_bbox)
    if rotate_vertical:
        image = image.rotate(90, expand=True)
    image_buffer = BytesIO()
    image.save(image_buffer, format="PNG")
    image_buffer.seek(0)
    return image_buffer


def get_merged_range_for_coordinate(worksheet: Any, coordinate: str) -> Any | None:
    for merged_range in worksheet.merged_cells.ranges:
        if coordinate in merged_range:
            return merged_range
    return None


def column_width_to_pixels(width: float | None) -> float:
    effective_width = 8.43 if width in (None, 0) else float(width)
    return max((effective_width * 7) + 5, 24)


def row_height_to_pixels(height: float | None) -> float:
    effective_height = 15 if height in (None, 0) else float(height)
    return max(effective_height * (96 / 72), 20)


def get_signature_bounds(worksheet: Any, coordinate: str) -> tuple[float, float]:
    merged_range = get_merged_range_for_coordinate(worksheet, coordinate)
    if merged_range is None:
        cell = worksheet[coordinate]
        column_width = worksheet.column_dimensions[get_column_letter(cell.column)].width
        row_height = worksheet.row_dimensions[cell.row].height
        return column_width_to_pixels(column_width), row_height_to_pixels(row_height)

    total_width = 0.0
    for column in range(merged_range.min_col, merged_range.max_col + 1):
        column_letter = get_column_letter(column)
        total_width += column_width_to_pixels(worksheet.column_dimensions[column_letter].width)

    total_height = 0.0
    for row in range(merged_range.min_row, merged_range.max_row + 1):
        total_height += row_height_to_pixels(worksheet.row_dimensions[row].height)

    return total_width, total_height


def get_signature_anchor_origin(worksheet: Any, coordinate: str) -> tuple[int, int]:
    merged_range = get_merged_range_for_coordinate(worksheet, coordinate)
    if merged_range is None:
        cell = worksheet[coordinate]
        return cell.column, cell.row
    return merged_range.min_col, merged_range.min_row


def fit_signature_dimensions(
    worksheet: Any,
    coordinate: str,
    original_width: float,
    original_height: float,
    requested_width: int,
    requested_height: int,
) -> tuple[int, int]:
    max_width, max_height = get_signature_bounds(worksheet, coordinate)
    max_width = max(min(max_width - 8, requested_width), 24)
    max_height = max(min(max_height - 4, requested_height), 16)

    if original_width <= 0 or original_height <= 0:
        return int(max_width), int(max_height)

    scale = min(max_width / original_width, max_height / original_height)
    scale = min(scale, 1.0)
    return max(int(original_width * scale), 1), max(int(original_height * scale), 1)


def build_signature_anchor(
    worksheet: Any,
    coordinate: str,
    width: int,
    height: int,
) -> OneCellAnchor:
    anchor_col, anchor_row = get_signature_anchor_origin(worksheet, coordinate)
    bounds_width, bounds_height = get_signature_bounds(worksheet, coordinate)
    offset_x = max(int((bounds_width - width) / 2), 0)
    offset_y = max(int((bounds_height - height) / 2), 0)
    marker = AnchorMarker(
        col=anchor_col - 1,
        colOff=pixels_to_EMU(offset_x),
        row=anchor_row - 1,
        rowOff=pixels_to_EMU(offset_y),
    )
    ext = XDRPositiveSize2D(pixels_to_EMU(width), pixels_to_EMU(height))
    return OneCellAnchor(_from=marker, ext=ext)


def write_signature_or_text_cell(
    worksheet: Any,
    coordinate: str,
    value: str,
    width: int,
    height: int,
) -> None:
    signature_candidate = find_signature_candidate(value)
    if signature_candidate is not None:
        image_source = get_signature_image_source(signature_candidate)
        if image_source is not None:
            add_signature_image(worksheet, coordinate, image_source, width=width, height=height)
            return
    write_template_cell(worksheet, coordinate, value)


def write_signature_or_text_slot(
    worksheet: Any,
    row: int,
    column: int,
    value: str,
    width: int,
    height: int,
) -> None:
    coordinate = worksheet.cell(row=row, column=column).coordinate
    signature_candidate = find_signature_candidate(value)
    if signature_candidate is not None:
        image_source = get_signature_image_source(signature_candidate)
        if image_source is not None:
            add_signature_image(
                worksheet,
                coordinate,
                image_source,
                width=width,
                height=height,
                rotate_vertical=True,
            )
            return
    write_slot_value(
        worksheet,
        row,
        column,
        value,
        font_size=18,
        rotate_like_hours=True,
    )


def write_signature_or_text_status(
    worksheet: Any,
    row: int,
    start_col: int,
    value: str,
    width: int,
    height: int,
) -> None:
    coordinate = worksheet.cell(row=row, column=start_col).coordinate
    signature_candidate = find_signature_candidate(value)
    if signature_candidate is not None:
        image_source = get_signature_image_source(signature_candidate)
        if image_source is not None:
            add_signature_image(worksheet, coordinate, image_source, width=width, height=height)
            return
    write_day_status(worksheet, row, start_col, value)


def write_day_status(worksheet: Any, row: int, start_col: int, value: str) -> None:
    cell = worksheet.cell(row=row, column=start_col)
    cell.value = value
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    cell.font = Font(size=20)


def write_slot_value(
    worksheet: Any,
    row: int,
    column: int,
    value: str,
    font_size: int = 9,
    rotate_like_hours: bool = False,
) -> None:
    cell = worksheet.cell(row=row, column=column)
    cell.value = value
    cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
        text_rotation=90 if rotate_like_hours else 0,
        wrap_text=False,
    )
    cell.font = Font(size=font_size)


def ensure_status_row_merges(worksheet: Any, payload: dict[str, Any]) -> None:
    rows_to_merge = get_form_definition(payload["metadata"]["form_key"])["layout"]["status_rows_to_merge"]
    for day in range(1, 32):
        start_col = DAY_BLOCK_START_COLUMNS[day]
        for row in rows_to_merge:
            cell_range = f"{worksheet.cell(row=row, column=start_col).coordinate}:{worksheet.cell(row=row, column=start_col + 2).coordinate}"
            if not is_range_already_merged(worksheet, cell_range):
                worksheet.merge_cells(cell_range)


def is_range_already_merged(worksheet: Any, cell_range: str) -> bool:
    return any(str(merged_range) == cell_range for merged_range in worksheet.merged_cells.ranges)


def calculate_corrected_temperature(
    raw_value: str,
    correction_bands: dict[str, Any],
    correction_factors: dict[str, Any],
    correction_operations: dict[str, str],
) -> str:
    measured = parse_measurement_number(raw_value)
    if measured is None:
        return ""

    normalized_bands: list[dict[str, Any]] = []
    for key, band in correction_bands.items():
        min_value = float(band["min"])
        max_value = float(band["max"])
        if max_value < min_value:
            reparsed_bounds = parse_range_bounds(str(band.get("label", "")))
            if reparsed_bounds is not None:
                min_value, max_value = reparsed_bounds
        lower_bound = min(min_value, max_value)
        upper_bound = max(min_value, max_value)
        normalized_bands.append(
            {
                "key": key,
                "min": lower_bound,
                "max": upper_bound,
            }
        )

    normalized_bands.sort(key=lambda item: (item["min"], item["max"]))

    factor_key = None
    for index, band in enumerate(normalized_bands):
        if band["min"] <= measured <= band["max"]:
            factor_key = band["key"]
            break

    if factor_key is None:
        for index in range(1, len(normalized_bands)):
            previous_band = normalized_bands[index - 1]
            current_band = normalized_bands[index]
            if previous_band["max"] < measured < current_band["min"]:
                factor_key = current_band["key"] if measured < 0 else previous_band["key"]
                break

    if factor_key is None or factor_key not in correction_factors:
        return format_decimal_value(measured)

    factor = float(correction_factors[factor_key])
    operation = correction_operations[factor_key]
    corrected = measured + factor if operation == "+" else measured - factor
    return format_decimal_value(corrected)


def maybe_autosave_payload(payload: dict[str, Any]) -> None:
    if not can_edit_daily_records():
        return

    last_saved_signature = st.session_state.get("last_saved_payload_signature")
    current_signature = get_payload_signature(payload)
    if last_saved_signature is None:
        remember_saved_snapshot(payload)
        return
    if current_signature == last_saved_signature:
        return

    now = time.time()
    last_attempt = float(st.session_state.get("last_autosave_attempt_at", 0.0))
    if now - last_attempt < AUTOSAVE_DEBOUNCE_SECONDS:
        return

    st.session_state["last_autosave_attempt_at"] = now
    try:
        saved_backend = save_payload(payload)
        st.session_state["last_autosave_at"] = get_local_now().strftime("%H:%M:%S")
        st.session_state["last_autosave_backend"] = saved_backend
        st.session_state["last_autosave_error"] = ""
    except Exception as exc:
        st.session_state["last_autosave_error"] = str(exc)


def render_actions(payload: dict[str, Any]) -> None:
    st.subheader("5. Guardado y exportacion")
    allow_daily_edits = can_edit_daily_records()
    allow_export = can_export_period()
    allow_reset = is_admin_role()
    maybe_autosave_payload(payload)
    errors = validate_payload(payload)
    if errors:
        st.warning("Hay datos pendientes antes de exportar.")
        for error in errors[:8]:
            st.write(f"- {error}")
        if len(errors) > 8:
            st.write(f"- ... y {len(errors) - 8} mas.")
    else:
        st.success("La captura esta completa para exportar la plantilla.")

    autosave_error = str(st.session_state.get("last_autosave_error", "")).strip()
    autosave_at = str(st.session_state.get("last_autosave_at", "")).strip()
    autosave_backend = str(st.session_state.get("last_autosave_backend", "")).strip()
    if autosave_error:
        st.caption(f"Autoguardado con problema: {autosave_error}")
    elif autosave_at:
        autosave_label = "respaldo local" if autosave_backend == "Local JSON" else "correcto"
        st.caption(f"Autoguardado {autosave_label}: {autosave_at}")

    action_columns_count = sum([allow_daily_edits, allow_export, allow_reset]) or 1
    action_columns = st.columns(action_columns_count)
    column_index = 0

    if allow_daily_edits:
        save_col = action_columns[column_index]
        column_index += 1
        save_label = "Guardar captura" if is_capture_role() else "Guardar borrador"
        if save_col.button(save_label, use_container_width=True):
            try:
                saved_backend = save_payload(payload)
                log_activity(
                    "guardar_borrador",
                    "Respaldo local" if saved_backend == "Local JSON" else "Guardado principal",
                    payload,
                )
                if saved_backend == "Local JSON":
                    st.warning("Se guardo un respaldo local del borrador.")
                else:
                    st.success("Se guardo la captura.")
            except Exception as exc:
                st.error(f"No se pudo guardar la captura: {exc}")

    if allow_export:
        export_col = action_columns[column_index]
        column_index += 1
        if export_col.button("Preparar Excel", use_container_width=True):
            if errors:
                st.error("Completa los campos pendientes antes de exportar.")
                return
            try:
                save_payload(payload)
            except Exception:
                pass
            excel_file = populate_template(payload)
            log_activity("exportar_excel", "Genero el archivo Excel", payload)
            form_definition = get_form_definition(payload["metadata"]["form_key"])
            form_code = form_definition["source_file"].split()[0].replace(".xlsx", "")
            filename = (
                f"{form_code}_{payload['metadata']['equipment_code']}_{payload['metadata']['year']}_{payload['metadata']['month']:02d}.xlsx"
            )
            st.download_button(
                "Descargar formato llenado",
                data=excel_file,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

    if allow_reset and action_columns[column_index].button("Limpiar periodo", use_container_width=True):
        log_activity("limpiar_periodo", "Restablecio el periodo actual", payload)
        current_period_key = get_period_key(payload)
        clear_period_widget_state(current_period_key)
        st.session_state.payload = build_default_payload(
            equipment_code=str(payload["metadata"]["equipment_code"]),
            form_key=str(payload["metadata"]["form_key"]),
        )
        st.session_state.payload["metadata"]["month"] = int(payload["metadata"]["month"])
        st.session_state.payload["metadata"]["year"] = int(payload["metadata"]["year"])
        st.session_state.period_key = get_period_key(st.session_state.payload)
        st.rerun()

    if not allow_daily_edits and not allow_export:
        st.caption("Tu perfil no tiene acciones disponibles para este periodo.")
    elif not allow_export:
        st.caption("Solo calidad o admin pueden cerrar y exportar el formato.")


MAIN_SECTIONS = [
    "Captura del periodo",
    "Reportes",
    "Lista maestra",
    "Trazabilidad y validacion",
]


def get_available_main_sections() -> list[str]:
    role = current_user_role()
    if role == "admin":
        return MAIN_SECTIONS
    if role == "calidad":
        return ["Captura del periodo", "Reportes", "Lista maestra", "Trazabilidad y validacion"]
    return ["Captura del periodo"]


def render_main_section_selector() -> str:
    available_sections = get_available_main_sections()
    previous_selection = str(st.session_state.get("main_section", available_sections[0]))
    if previous_selection not in available_sections:
        st.session_state["main_section"] = available_sections[0]

    if hasattr(st, "segmented_control"):
        return st.segmented_control(
            "Apartado",
            options=available_sections,
            default=available_sections[0],
            key="main_section",
            label_visibility="collapsed",
        ) or available_sections[0]
    return st.radio(
        "Apartado",
        options=available_sections,
        horizontal=True,
        key="main_section",
        label_visibility="collapsed",
    )


def main() -> None:
    form_keys = list(FORM_DEFINITIONS.keys())

    st.set_page_config(
        page_title="Formularios Digitales",
        layout="wide",
    )
    configure_users_backend()
    configure_storage_backend()
    configure_microsoft_backend()
    initialize_auth_state()

    if not st.session_state["autenticado"]:
        render_auth_screen()
        st.stop()

    for form_key in form_keys:
        definition = get_form_definition(form_key)
        if not template_source_available(form_key):
            st.error(f"No se encontro la plantilla {definition['source_file']}.")
            st.stop()

    if "payload" not in st.session_state:
        st.session_state.payload = load_saved_payload(
            form_key=DEFAULT_FORM_KEY,
            equipment_code=FORM_DEFINITIONS[DEFAULT_FORM_KEY]["default_equipment"],
        )
        remember_saved_snapshot(st.session_state.payload)

    payload = st.session_state.payload
    st.session_state.payload = hydrate_payload_corrections(payload)
    payload = st.session_state.payload
    previous_period_key = st.session_state.get("period_key", get_period_key(payload))
    equipment_configs = load_equipment_configs(payload["metadata"]["form_key"])
    equipment_codes = list(equipment_configs.keys())

    st.title("Digitalizacion de formatos")
    st.caption(
        f"Captura guiada y exportacion automatica para {payload['metadata']['form_label']} en {payload['metadata']['equipment_code']}."
    )

    selected_form_key, selected_equipment = render_sidebar(payload, form_keys, equipment_codes)
    if selected_form_key != payload["metadata"]["form_key"]:
        target_equipment = FORM_DEFINITIONS[selected_form_key]["default_equipment"]
        st.session_state.payload = load_saved_payload(
            form_key=selected_form_key,
            equipment_code=target_equipment,
            year=int(payload["metadata"]["year"]),
            month=int(payload["metadata"]["month"]),
        )
        remember_saved_snapshot(st.session_state.payload)
        st.session_state.period_key = get_period_key(st.session_state.payload)
        st.rerun()

    if selected_equipment != payload["metadata"]["equipment_code"]:
        target_period_key = (
            f"{payload['metadata']['form_key']}_{selected_equipment}_{int(payload['metadata']['year'])}_{int(payload['metadata']['month']):02d}"
        )
        clear_period_widget_state(target_period_key)
        st.session_state.payload = load_saved_payload(
            form_key=str(payload["metadata"]["form_key"]),
            equipment_code=selected_equipment,
            year=int(payload["metadata"]["year"]),
            month=int(payload["metadata"]["month"]),
        )
        remember_saved_snapshot(st.session_state.payload)
        st.session_state.period_key = get_period_key(st.session_state.payload)
        st.rerun()

    selected_section = render_main_section_selector()

    if selected_section == "Captura del periodo":
        payload = st.session_state.payload
        render_configuration(payload)

        current_period_key = get_period_key(payload)
        if current_period_key != previous_period_key:
            clear_period_widget_state(current_period_key)
            st.session_state.payload = load_saved_payload(
                form_key=str(payload["metadata"]["form_key"]),
                equipment_code=str(payload["metadata"]["equipment_code"]),
                year=int(payload["metadata"]["year"]),
                month=int(payload["metadata"]["month"]),
            )
            remember_saved_snapshot(st.session_state.payload)
            st.session_state.period_key = current_period_key
            st.rerun()

        st.session_state.period_key = current_period_key
        payload = st.session_state.payload

        render_non_working_days(payload)
        render_daily_capture(payload)
        render_monthly_closure(payload)
        render_actions(payload)

    payload = st.session_state.payload
    if selected_section == "Reportes":
        render_reports(payload)

    if selected_section == "Lista maestra":
        render_master_list()

    if selected_section == "Trazabilidad y validacion":
        render_traceability_and_validation(payload)


if __name__ == "__main__":
    main()
